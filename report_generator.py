"""
Sistema de Generación de Reportes Profesionales
VERSIÓN CORREGIDA - Genera reportes PDF y Excel con diseño elegante y corporativo.

CORRECCIONES APLICADAS:
1. Fix en generación de PDF con xhtml2pdf (uso correcto de pisa.CreatePDF)
2. Fix en generación de Excel para evitar datos corruptos
3. Mejor manejo de errores y encoding
"""

import os
from db_helper import get_db_connection, is_sqlite, is_mysql, execute_query, get_date_format, get_year_function, get_month_function, get_day_function, get_current_timestamp, get_date_sub, get_date_add, get_coalesce_function
import json
from datetime import datetime
from io import BytesIO
import base64

# Importar bibliotecas para PDF (xhtml2pdf) y Excel
try:
    from xhtml2pdf import pisa  # <- CORRECCIÓN: Importar pisa en lugar de pml/pmlDoc
    XHTML2PDF_DISPONIBLE = True
except ImportError:
    XHTML2PDF_DISPONIBLE = False
    print("xhtml2pdf no disponible. Instalar con: pip install xhtml2pdf")

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    PANDAS_DISPONIBLE = True
except ImportError:
    PANDAS_DISPONIBLE = False
    print("Pandas/OpenPyXL no disponible. Instalar con: pip install pandas openpyxl")


class ReporteGenerator:
    """Generador de reportes profesionales"""
    
    # Paleta de colores corporativa
    COLORES = {
        'primario': '#1E3A8A',      # Azul Marino - Institucional
        'secundario': '#3B82F6',    # Azul Brillante - Acciones
        'acento': '#10B981',        # Verde Esmeralda - Éxito/Dinero
        'fondo': '#FFFFFF',         # Blanco puro
        'texto': '#1F2937',         # Gris oscuro - Legibilidad
        'texto_claro': '#6B7280',   # Gris medio
        'borde': '#E5E7EB',         # Gris claro - Bordes
        'header_bg': '#F3F4F6',     # Gris muy claro - Headers
        'warning': '#F59E0B',       # Naranja - Advertencia
        'danger': '#EF4444',        # Rojo - Error
    }
    
    def __init__(self, db_path=None):
        # Determinar el directorio base de la aplicación
        self.base_dir = None
        try:
            # Intentar importar Flask para obtener app.root_path
            from flask import current_app
            self.base_dir = current_app.root_path
        except (ImportError, RuntimeError):
            # Si no hay Flask, usar el directorio del archivo report_generator.py
            self.base_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Si no se proporciona db_path, buscar en ubicaciones estándar
        if db_path is None:
            # Buscar la base de datos en varias ubicaciones
            db_candidates = [
                os.path.join(self.base_dir, 'sistema.db'),
                os.path.join(self.base_dir, '..', 'sistema.db'),
                'sistema.db',
            ]
            for candidate in db_candidates:
                if os.path.exists(candidate):
                    db_path = os.path.abspath(candidate)
                    break
            else:
                db_path = 'sistema.db'  # Usar relativo como último recurso
        
        self.db_path = db_path
        self.config = None
        self.cargar_configuracion()
        print(f"[DEBUG] ReporteGenerator inicializado con db_path: {self.db_path}")
        print(f"[DEBUG] Base directory: {self.base_dir}")
    
    def _get_connection(self):
        """Obtiene conexión a la base de datos"""
        conn = get_db_connection()
        # db_helper ya configura row_factory para SQLite o devuelve dict para MySQL
        return conn
    
    def cargar_configuracion(self):
        """Carga la configuración de la empresa"""
        config = execute_query('SELECT * FROM configuraciones ORDER BY id DESC LIMIT 1', fetch=True)
        if config:
            config = config[0] # execute_query devuelve una lista de diccionarios
            print(f"[DEBUG] Config keys: {list(config.keys())}")
            print(f"[DEBUG] empresa_logo value: {repr(config.get('empresa_logo'))}")
            print(f"[DEBUG] color_iconos value: {repr(config.get('color_iconos'))}")
        
        def safe_color(value, default):
            """Valida que el valor sea un color hex válido, sino usa el default"""
            if not value:
                return default
            v = str(value).strip()
            # Debe empezar con # y tener 4 o 7 caracteres, o ser un nombre CSS básico
            if v.startswith('#') and len(v) in (4, 7):
                return v
            # Si parece un nombre de columna o valor inválido, usar default
            return default

        if config:
            self.config = {
                'empresa_nombre': config.get('empresa_nombre', ''),
                'empresa_logo': config.get('empresa_logo'),
                'color_primario': safe_color(config.get('color_botones'), self.COLORES['primario']),
                'color_secundario': safe_color(config.get('color_botones_secundarios'), self.COLORES['secundario']),
                'color_acento': safe_color(config.get('color_iconos'), self.COLORES['acento']),
            }
        else:
            self.config = {
                'empresa_nombre': '',
                'empresa_logo': None,
                'color_primario': self.COLORES['primario'],
                'color_secundario': self.COLORES['secundario'],
                'color_acento': self.COLORES['acento'],
            }
    
    def get_logo_path(self):
        """Obtiene la ruta absoluta del archivo de logo"""
        if not self.config['empresa_logo']:
            print("[DEBUG] No hay empresa_logo configurado")
            return None
        
        # Usar el base_dir guardado en __init__
        base_dir = getattr(self, 'base_dir', None)
        if not base_dir:
            try:
                from flask import current_app
                base_dir = current_app.root_path
            except (ImportError, RuntimeError):
                base_dir = os.path.dirname(os.path.abspath(__file__))
        
        print(f"[DEBUG] get_logo_path - base_dir: {base_dir}")
        print(f"[DEBUG] get_logo_path - logo configurado: {self.config['empresa_logo']}")
        
        # Buscar el logo en múltiples ubicaciones posibles
        posibles_rutas = [
            os.path.join(base_dir, 'static', 'uploads', self.config['empresa_logo']),
            os.path.join(base_dir, 'static', 'uploads', 'perfiles', self.config['empresa_logo']),
            os.path.join(base_dir, 'uploads', self.config['empresa_logo']),
            os.path.join(base_dir, '..', 'static', 'uploads', self.config['empresa_logo']),
            os.path.join(base_dir, '..', 'static', 'uploads', 'perfiles', self.config['empresa_logo']),
            # Rutas relativas desde el directorio actual de trabajo
            os.path.join('static', 'uploads', self.config['empresa_logo']),
            os.path.join('static', 'uploads', 'perfiles', self.config['empresa_logo']),
            os.path.join('uploads', self.config['empresa_logo']),
            # Ruta directa si es una ruta absoluta
            self.config['empresa_logo'],
        ]
        
        logo_path = None
        for ruta in posibles_rutas:
            # Normalizar la Ruta y verificar si existe
            ruta_normalizada = os.path.normpath(ruta)
            if os.path.exists(ruta_normalizada):
                # Convertir a ruta absoluta
                logo_path = os.path.abspath(ruta_normalizada)
                print(f"[DEBUG] Logo encontrado en: {logo_path}")
                break
        
        if not logo_path:
            print(f"[DEBUG] Logo no encontrado. Buscado: {self.config['empresa_logo']}")
            print(f"[DEBUG] Rutas intentadas: {[os.path.normpath(r) for r in posibles_rutas]}")
        
        return logo_path
    
    def get_logo_base64(self):
        """Obtiene el logo como base64 para embeber en PDFs.
        Corrige automáticamente la orientación EXIF para evitar que la imagen
        aparezca rotada en el PDF (problema común en fotos tomadas con celular).
        """
        if not self.config['empresa_logo']:
            return None
        
        # Usar base_dir igual que get_logo_path para rutas absolutas
        base_dir = getattr(self, 'base_dir', None)
        if not base_dir:
            try:
                from flask import current_app
                base_dir = current_app.root_path
            except (ImportError, RuntimeError):
                base_dir = os.path.dirname(os.path.abspath(__file__))

        logo_nombre = self.config['empresa_logo']

        posibles_rutas = [
            os.path.join(base_dir, 'static', 'uploads', logo_nombre),
            os.path.join(base_dir, 'static', 'uploads', 'perfiles', logo_nombre),
            os.path.join(base_dir, 'uploads', logo_nombre),
            os.path.join('static', 'uploads', logo_nombre),
            os.path.join('static', 'uploads', 'perfiles', logo_nombre),
            logo_nombre,
        ]
        
        logo_path = None
        for ruta in posibles_rutas:
            ruta_norm = os.path.normpath(ruta)
            if os.path.exists(ruta_norm):
                logo_path = os.path.abspath(ruta_norm)
                print(f"[DEBUG] Logo base64 encontrado en: {logo_path}")
                break
        
        if not logo_path:
            print(f"Logo no encontrado en: {logo_nombre}")
            return None
        
        try:
            # Intentar corregir orientación EXIF con Pillow
            try:
                from PIL import Image, ExifTags
                import io as _io

                img = Image.open(logo_path)

                # Buscar el tag de orientación EXIF
                exif = None
                try:
                    exif = img._getexif()
                except Exception:
                    exif = None

                if exif:
                    orientation_key = next(
                        (k for k, v in ExifTags.TAGS.items() if v == 'Orientation'), None
                    )
                    if orientation_key and orientation_key in exif:
                        orientation = exif[orientation_key]
                        # Aplicar la rotación según el valor EXIF
                        rotaciones = {
                            3: Image.ROTATE_180,
                            6: Image.ROTATE_270,
                            8: Image.ROTATE_90,
                        }
                        if orientation in rotaciones:
                            img = img.transpose(rotaciones[orientation])
                            print(f"[DEBUG] Logo: orientación EXIF {orientation} corregida")

                # Convertir a PNG en memoria para máxima compatibilidad
                buffer = _io.BytesIO()
                # Convertir a RGB si tiene canal alpha (RGBA) para evitar problemas con JPEG
                if img.mode in ('RGBA', 'P', 'LA'):
                    img = img.convert('RGBA')
                    img.save(buffer, format='PNG')
                    logo_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
                    return f"data:image/png;base64,{logo_data}"
                else:
                    img = img.convert('RGB')
                    img.save(buffer, format='JPEG', quality=95)
                    logo_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
                    return f"data:image/jpeg;base64,{logo_data}"

            except ImportError:
                # Pillow no disponible: leer el archivo directamente (sin corrección EXIF)
                print("[AVISO] Pillow no instalado. La imagen puede aparecer rotada. "
                      "Instala con: pip install Pillow")
                with open(logo_path, 'rb') as f:
                    logo_data = base64.b64encode(f.read()).decode('utf-8')
                ext = os.path.splitext(logo_path)[1].lower()
                if ext in ['.jpg', '.jpeg']:
                    return f"data:image/jpeg;base64,{logo_data}"
                elif ext in ['.png']:
                    return f"data:image/png;base64,{logo_data}"
                elif ext in ['.gif']:
                    return f"data:image/gif;base64,{logo_data}"
                else:
                    return f"data:image/png;base64,{logo_data}"

        except Exception as e:
            print(f"Error cargando logo: {str(e)}")
            return None
    
    
    
    def generar_pdf(self, html_content, filename="reporte.pdf"):
        """
        Genera un PDF desde el contenido HTML usando xhtml2pdf
        CON MEJOR MANEJO DE IMÁGENES
        """
        if not XHTML2PDF_DISPONIBLE:
            raise Exception("xhtml2pdf no está disponible. Por favor instale: pip install xhtml2pdf")
        
        try:
            # Crear buffer para el PDF
            pdf_buffer = BytesIO()
            
            # Configuración adicional para xhtml2pdf
            context = {}
            
            # Crear función callback para logging
            def make_pdf(html):
                return pisa.CreatePDF(
                    src=html.encode('utf-8'),
                    dest=pdf_buffer,
                    encoding='utf-8',
                    show_error_as_pdf=True
                )
            
            # Convertir HTML a PDF
            pisa_status = make_pdf(html_content)
            
            # Verificar si hubo errores
            if pisa_status.err:
                print(f"Advertencias al generar PDF: {pisa_status.warn}")
                
                # Intentar una segunda versión sin logo si falla
                if "Image" in str(pisa_status.err):
                    print("Problema con imagen, intentando sin logo...")
                    # Remover la sección del logo del HTML
                    html_simple = html_content.replace('<img', '<!-- <img').replace('/>', '/> -->')
                    pdf_buffer = BytesIO()
                    pisa_status = make_pdf(html_simple)
            
            if pisa_status.err:
                raise Exception(f"Error al generar PDF: {pisa_status.err}")
            
            # Mover el puntero al inicio del buffer
            pdf_buffer.seek(0)
            
            return pdf_buffer
            
        except Exception as e:
            raise Exception(f"Error generando PDF: {str(e)}")
    
    def _agregar_encabezado_excel(self, ws, sheet_name, incluir_logo=True):
        """
        Función auxiliar para agregar el encabezado con logo a cualquier hoja de Excel.
        Returns:
            int: El número de fila donde deben comenzar los datos de la tabla
        """
        from openpyxl.drawing.image import Image
        
        # Convertir colores
        color_primario = self.config['color_primario'].replace('#', '')
        color_secundario = self.config['color_secundario'].replace('#', '')
        
        start_row = 1
        
        if incluir_logo:
            # NUEVO: Usar ruta directa del archivo (igual que en PHP)
            logo_path = self.get_logo_path()
            if logo_path and os.path.exists(logo_path):
                try:
                    # Agregar logo a Excel usando ruta directa
                    img = Image(logo_path)
                    img.width = 180  # Ancho en píxeles
                    img.height = 60  # Alto en píxeles
                    
                    # Posicionar logo en celda D1 (más a la derecha)
                    ws.add_image(img, 'B2')
                    
                    # Ajustar altura de fila para el logo
                    ws.row_dimensions[1].height = 50
                    ws.row_dimensions[2].height = 20
                    
                    # Subtítulo solo (sin nombre de empresa)
                    ws.merge_cells('D3:G3')
                    subtitulo_cell = ws['D3']
                    subtitulo_cell.value = "Reportes Profesionales"
                    subtitulo_cell.font = Font(size=12, italic=True, color=color_secundario)
                    subtitulo_cell.alignment = Alignment(horizontal='center')
                    
                    start_row = 6
                    
                    print(f"[DEBUG] Logo agregado exitosamente desde: {logo_path}")
                    
                except Exception as e:
                    print(f"⚠️ No se pudo agregar logo a Excel: {e}")
                    # Si falla el logo, crear encabezado simple
                    ws.merge_cells('A1:G2')
                    titulo_cell = ws['A1']
                    titulo_cell.value = self.config.get('empresa_nombre', '')
                    titulo_cell.font = Font(size=18, bold=True, color="FFFFFF")
                    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
                    titulo_cell.fill = PatternFill(start_color=color_primario, 
                                                    end_color=color_primario, 
                                                    fill_type='solid')
                    
                    ws['A3'] = f"Reporte: {sheet_name}"
                    ws['A3'].font = Font(bold=True)
                    
                    ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                    ws['A4'].font = Font(size=10)
                    
                    start_row = 6
            else:
                # No hay logo disponible, crear encabezado simple
                print("[DEBUG] Logo no encontrado, creando encabezado sin logo")
                ws.merge_cells('A1:G2')
                titulo_cell = ws['A1']
                titulo_cell.value = self.config.get('empresa_nombre', '')
                titulo_cell.font = Font(size=18, bold=True, color="FFFFFF")
                titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
                titulo_cell.fill = PatternFill(start_color=color_primario, 
                                                end_color=color_primario, 
                                                fill_type='solid')
                
                ws['A3'] = f"Reporte: {sheet_name}"
                ws['A3'].font = Font(bold=True)
                
                ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                ws['A4'].font = Font(size=10)
                
                start_row = 6
        else:
            # Sin logo pero con encabezado
            ws.merge_cells('A1:G2')
            titulo_cell = ws['A1']
            titulo_cell.value = self.config.get('empresa_nombre', '')
            titulo_cell.font = Font(size=18, bold=True, color="FFFFFF")
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            titulo_cell.fill = PatternFill(start_color=color_primario, 
                                            end_color=color_primario, 
                                            fill_type='solid')
            
            ws['A3'] = f"Reporte: {sheet_name}"
            ws['A3'].font = Font(bold=True)
            
            ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A4'].font = Font(size=10)
            
            start_row = 6
        
        return start_row
    
    # ============================================
    # GENERACIÓN DE EXCEL - VERSIÓN CORREGIDA
    # ============================================
    
    def generar_excel(self, data, headers, filename="reporte.xlsx", sheet_name="Reporte", incluir_logo=True):
        """
        Genera un Excel profesional desde los datos CON LOGO
        CORRECCIÓN: Mejor manejo de datos para evitar corrupción
        """
        if not PANDAS_DISPONIBLE:
            raise Exception("Pandas/OpenPyXL no está disponible. Por favor instale: pip install pandas openpyxl")
        
        try:
            from openpyxl.drawing.image import Image
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel limita nombres a 31 caracteres
            
            # Convertir color hexadecimal a formato Excel (sin #)
            color_primario = self.config['color_primario'].replace('#', '')
            color_secundario = self.config['color_secundario'].replace('#', '')
            color_borde = self.COLORES['borde'].replace('#', '')
            color_header_bg = self.COLORES['header_bg'].replace('#', '')
            
            # ============================================
            # SECCIÓN 1: AGREGAR LOGO Y ENCABEZADO
            # ============================================
            start_row = 1
            
            if incluir_logo:
                # NUEVO: Usar ruta directa del archivo (igual que en PHP)
                logo_path = self.get_logo_path()
                if logo_path and os.path.exists(logo_path):
                    try:
                        # Agregar logo a Excel usando ruta directa
                        img = Image(logo_path)
                        img.width = 180  # Ancho en píxeles
                        img.height = 60  # Alto en píxeles
                        
                        # Posicionar logo en celda A1
                        ws.add_image(img, 'A1')
                        
                        # Ajustar altura de fila para el logo
                        ws.row_dimensions[1].height = 50
                        ws.row_dimensions[2].height = 20
                        
                        # Información de la empresa al lado del logo
                        ws.merge_cells('C1:F2')
                        empresa_cell = ws['C1']
                        empresa_cell.value = self.config.get('empresa_nombre', '')
                        empresa_cell.font = Font(size=24, bold=True, color=color_primario)
                        empresa_cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Subtítulo
                        ws.merge_cells('C3:F3')
                        subtitulo_cell = ws['C3']
                        subtitulo_cell.value = "Sistema de Gestión Deportiva"
                        subtitulo_cell.font = Font(size=12, italic=True, color=color_secundario)
                        subtitulo_cell.alignment = Alignment(horizontal='center')
                        
                        start_row = 6
                        
                    except Exception as e:
                        print(f"⚠️ No se pudo agregar logo a Excel: {e}")
                        # Si falla el logo, crear encabezado simple
                        ws.merge_cells('A1:G2')
                        titulo_cell = ws['A1']
                        titulo_cell.value = self.config.get('empresa_nombre', '')
                        titulo_cell.font = Font(size=18, bold=True, color="FFFFFF")
                        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
                        titulo_cell.fill = PatternFill(start_color=color_primario, 
                                                        end_color=color_primario, 
                                                        fill_type='solid')
                        
                        ws['A3'] = f"Reporte: {sheet_name}"
                        ws['A3'].font = Font(bold=True)
                        
                        ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                        ws['A4'].font = Font(size=10)
                        
                        start_row = 6
                else:
                    # No hay logo disponible, crear encabezado simple
                    ws.merge_cells('A1:G2')
                    titulo_cell = ws['A1']
                    titulo_cell.value = self.config.get('empresa_nombre', '')
                    titulo_cell.font = Font(size=18, bold=True, color="FFFFFF")
                    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
                    titulo_cell.fill = PatternFill(start_color=color_primario, 
                                                    end_color=color_primario, 
                                                    fill_type='solid')
                    
                    ws['A3'] = f"Reporte: {sheet_name}"
                    ws['A3'].font = Font(bold=True)
                    
                    ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                    ws['A4'].font = Font(size=10)
                    
                    start_row = 6
            else:
                # Sin logo pero con encabezado
                ws.merge_cells('A1:G2')
                titulo_cell = ws['A1']
                titulo_cell.value = self.config.get('empresa_nombre', '')
                titulo_cell.font = Font(size=18, bold=True, color="FFFFFF")
                titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
                titulo_cell.fill = PatternFill(start_color=color_primario, 
                                                end_color=color_primario, 
                                                fill_type='solid')
                
                ws['A3'] = f"Reporte: {sheet_name}"
                ws['A3'].font = Font(bold=True)
                
                ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                ws['A4'].font = Font(size=10)
                
                start_row = 6
            
            # ============================================
            # SECCIÓN 2: ENCABEZADOS DE LA TABLA
            # ============================================
            header_row = start_row
            
            # Estilos para encabezados
            header_fill = PatternFill(start_color=color_primario, 
                                    end_color=color_primario, 
                                    fill_type='solid')
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style='thin', color=color_borde),
                right=Side(style='thin', color=color_borde),
                top=Side(style='thin', color=color_borde),
                bottom=Side(style='thin', color=color_borde)
            )
            
            # Escribir encabezados
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = str(header)  # Convertir a string
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ============================================
            # SECCIÓN 3: DATOS DE LA TABLA
            # ============================================
            data_font = Font(size=10)
            alt_fill = PatternFill(start_color=color_header_bg, 
                                    end_color=color_header_bg, 
                                    fill_type='solid')
            
            # Escribir datos
            for row_num, row_data in enumerate(data, header_row + 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # Convertir valor de forma segura
                    if value is None:
                        cell.value = ""
                    elif isinstance(value, (int, float)):
                        cell.value = value
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                    elif isinstance(value, datetime):
                        cell.value = value.strftime('%d/%m/%Y %H:%M')
                    else:
                        cell.value = str(value)
                    
                    # Aplicar estilos
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Filas alternas con color
                    if (row_num - header_row) % 2 == 0:
                        cell.fill = alt_fill
            
            # ============================================
            # SECCIÓN 4: PIE DE PÁGINA
            # ============================================
            footer_row = header_row + len(data) + 2
            
            # Línea separadora
            ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
            separator_cell = ws[f'A{footer_row}']
            separator_cell.border = Border(bottom=Side(style='medium', color=color_primario))
            
            # Información del pie
            footer_row += 1
            ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = f"Reporte generado - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            footer_cell.font = Font(size=9, italic=True, color=color_secundario)
            footer_cell.alignment = Alignment(horizontal='center')
            
            # Totales si aplica
            if len(data) > 0 and any(isinstance(cell, (int, float)) for row in data for cell in row):
                footer_row += 1
                total_cell = ws.cell(row=footer_row, column=len(headers)-1)
                total_cell.value = "Total Registros:"
                total_cell.font = Font(bold=True)
                total_cell.alignment = Alignment(horizontal='right')
                
                total_val_cell = ws.cell(row=footer_row, column=len(headers))
                total_val_cell.value = len(data)
                total_val_cell.font = Font(bold=True)
            
            # ============================================
            # SECCIÓN 5: AJUSTAR FORMATO
            # ============================================
            # Ajustar ancho de columnas automáticamente
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Saltar columnas fusionadas
                if column_letter == 'C' and incluir_logo:
                    ws.column_dimensions['A'].width = 40  # Ancho fijo para columna del logo
                    ws.column_dimensions['B'].width = 5   # Espacio entre logo y texto
                    continue
                
                for cell in column:
                    try:
                        if cell.value:
                            # Calcular longitud del texto
                            cell_length = len(str(cell.value))
                            
                            # Ajustar para fechas y números
                            if isinstance(cell.value, datetime) or 'fecha' in str(cell.value).lower():
                                cell_length = 12  # Fecha fija
                            elif isinstance(cell.value, (int, float)):
                                cell_length = min(15, cell_length + 5)
                            
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                # Ajustar ancho con límites
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar paneles (encabezados fijos)
            ws.freeze_panes = ws[f'A{header_row + 1}']
            
            # Ajustar altura de filas
            for row in range(header_row, header_row + len(data) + 1):
                ws.row_dimensions[row].height = 20
            
            # ============================================
            # SECCIÓN 6: GUARDAR ARCHIVO
            # ============================================
            # Guardar en buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except ImportError as e:
            raise Exception(f"Falta dependencia para Excel: {str(e)}")
        except Exception as e:
            raise Exception(f"Error generando Excel: {str(e)}")
    
    
    
    
    def generar_css_pdf(self, landscape=False):
        """Genera el CSS para los PDFs - SOLO REDUCE ESPACIO LOGO-REPORTE"""
        size = 'A4 landscape' if landscape else 'A4'
        margin = '10mm' if landscape else '10mm'
        
        return f"""
            @page {{
                size: {size};
                margin: {margin};
            }}
            
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            
            body {{
                font-family: Arial, Helvetica, sans-serif;
                font-size: 9pt;
                color: {self.COLORES['texto']};
                line-height: 1.4;
                margin: 0;
                padding: 0;
            }}
            
            /* HEADER - SOLO REDUCE ESPACIO VERTICAL INTERNO */
            .header-container {{
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                padding: 5px 15px 8px 15px;  /* REDUCIDO padding vertical (antes 10px 15px) */
                background: linear-gradient(135deg, {self.config['color_primario']} 0%, {self.config['color_secundario']} 100%);
                color: #000000;
                margin-bottom: 10px;  /* MANTENIDO igual */
                page-break-inside: avoid;
                border-bottom: 3px solid {self.config['color_acento']};
                text-align: center;
            }}
            
            /* LOGO - SIN ESPACIO ABAJO */
            .logo-section {{
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                margin-bottom: 0px;  /* REDUCIDO A CERO (antes 5px) */
            }}
            
            /* LOGO MÁS GRANDE Y SIN MARGEN ABAJO */
            .logo-section img {{
                height: 140px;  /* AUMENTADO A 140px */
                width: auto;
                max-width: 400px;
                object-fit: contain;
                display: block;
                margin: 0 auto 0px auto;  /* CERO margin abajo (antes varios px) */
                border-radius: 8px;
                background-color: white;
                padding: 5px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }}
            
            .logo-placeholder {{
                width: 400px;
                height: 140px;
                background-color: white;
                border-radius: 8px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: bold;
                color: {self.config['color_primario']};
                font-size: 36px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                text-align: center;
                padding: 10px;
                margin: 0 auto 0px auto;
            }}
            
            /* INFORMACIÓN DE EMPRESA - CON MENOS ESPACIO ARRIBA */
            .company-info {{
                text-align: center;
                margin-bottom: 4px;  /* MANTENIDO similar */
                margin-top: 0px;  /* SIN espacio arriba (antes tenía) */
            }}
            
            .company-info h1 {{
                font-size: 26pt;
                font-weight: bold;
                margin: 2px 0 3px 0;  /* REDUCIDO margen superior (antes más) */
                text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.2);
                line-height: 1.2;
            }}
            
            .company-info p {{
                font-size: 12pt;
                opacity: 0.9;
                font-style: italic;
                margin: 0 0 4px 0;  /* MANTENIDO normal */
                line-height: 1.2;
            }}
            
            /* METADATOS - CON MENOS ESPACIO ARRIBA */
            .report-meta {{
                text-align: center;
                font-size: 9pt;
                background-color: rgba(255, 255, 255, 0.1);
                padding: 6px 15px 8px 15px;
                border-radius: 6px;
                margin: 2px auto 0 auto;  /* REDUCIDO margen superior (antes 5px) */
                max-width: 90%;
            }}
            
            .report-meta p {{
                margin: 3px 0;  /* MANTENIDO igual */
                line-height: 1.2;
            }}
            
            /* CONTENIDO PRINCIPAL - MANTENIDO IGUAL */
            .content {{
                padding: 15px 15px 15px 15px;  /* MANTENIDO igual */
                min-height: calc(100vh - 180px);
            }}
            
            /* TÍTULO DEL REPORTE - MANTENIDO IGUAL */
            .report-title {{
                font-size: 16pt;
                font-weight: bold;
                color: {self.config['color_primario']};
                margin: 15px 0 10px 0;  /* MANTENIDO igual */
                padding-bottom: 6px;
                border-bottom: 2px solid {self.config['color_acento']};
                text-align: center;
            }}
            
            .report-subtitle {{
                font-size: 11pt;
                color: {self.COLORES['texto_claro']};
                margin-bottom: 12px;
                text-align: center;
            }}
            
            /* ESTADÍSTICAS - DISEÑO MEJORADO CON TABLA */
            .stats-table {{
                width: 100%;
                margin: 15px 0 20px 0;
                background-color: {self.config['color_primario']};
                border-radius: 8px;
                border-collapse: separate;
                border-spacing: 0;
            }}
            
            .stat-cell {{
                width: 33.33%;
                text-align: center;
                padding: 18px 15px;
                color: #FFFFFF;
                vertical-align: middle;
            }}
            
            .stat-cell-wide {{
                width: 50%;
                text-align: center;
                padding: 18px 15px;
                color: #FFFFFF;
                vertical-align: middle;
                border-right: 2px solid rgba(255, 255, 255, 0.3);
            }}
            
            .stat-cell-wide:last-child {{
                border-right: none;
            }}
            
            .stat-cell-middle {{
                border-left: 2px solid rgba(255, 255, 255, 0.3);
                border-right: 2px solid rgba(255, 255, 255, 0.3);
            }}
            
            .stat-label {{
                font-size: 9pt;
                color: #FFFFFF;
                margin-bottom: 8px;
                font-weight: bold;
            }}
            
            .stat-value {{
                font-size: 24pt;
                font-weight: bold;
                color: #FFFFFF;
            }}
            
            /* TABLA - MANTENIDO EXACTAMENTE IGUAL */
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 10px 0 15px 0;  /* MANTENIDO IGUAL */
                font-size: 9pt;
            }}
            
            table th {{
                background: {self.config['color_primario']};
                color: white;
                padding: 8px 10px;  /* MANTENIDO IGUAL */
                text-align: left;
                font-weight: bold;
                font-size: 9pt;
            }}
            
            table td {{
                padding: 7px 10px;  /* MANTENIDO IGUAL */
                border-bottom: 1px solid {self.COLORES['borde']};
                font-size: 9pt;
            }}
            
            table tr:nth-child(even) {{
                background: {self.COLORES['header_bg']};
            }}
            
            table tr:hover {{
                background: rgba(30, 58, 138, 0.05);
            }}
            
            /* ANCHOS DE COLUMNA ESPECÍFICOS PARA REPORTE DE MEMBRESÍAS */
            table.tabla-membresias {{ width: 100%; table-layout: fixed; }}
            table.tabla-membresias th,
            table.tabla-membresias td {{
                padding: 8px 6px;
                font-size: 9pt;
                text-align: center;
                vertical-align: middle;
                line-height: 1.4;
            }}
            table.tabla-membresias th {{
                font-size: 9pt;
                font-weight: bold;
                white-space: normal;
            }}
            table.tabla-membresias th:nth-child(1),
            table.tabla-membresias td:nth-child(1) {{ width: 25px; }}  /* N° */
            table.tabla-membresias th:nth-child(2),
            table.tabla-membresias td:nth-child(2) {{ width: 50px; }}  /* Codigo */
            table.tabla-membresias th:nth-child(3),
            table.tabla-membresias td:nth-child(3) {{ width: 120px; text-align: left; }}  /* Nombre */
            table.tabla-membresias th:nth-child(4),
            table.tabla-membresias td:nth-child(4) {{ width: 50px; }}  /* Precio */
            table.tabla-membresias th:nth-child(5),
            table.tabla-membresias td:nth-child(5) {{ width: 50px; }}  /* Duracion */
            table.tabla-membresias th:nth-child(6),
            table.tabla-membresias td:nth-child(6) {{ width: 75px; }}  /* Fecha Creacion */
            table.tabla-membresias th:nth-child(7),
            table.tabla-membresias td:nth-child(7) {{ width: 45px; }}  /* Tiene QR */
            table.tabla-membresias th:nth-child(8),
            table.tabla-membresias td:nth-child(8) {{ width: 90px; }}  /* Permite Aplazamiento */
            table.tabla-membresias th:nth-child(9),
            table.tabla-membresias td:nth-child(9) {{ width: 90px; }}  /* Permite Invitados */
            table.tabla-membresias th:nth-child(10),
            table.tabla-membresias td:nth-child(10) {{ width: 80px; }}  /* Numero Invitados */
            table.tabla-membresias th:nth-child(11),
            table.tabla-membresias td:nth-child(11) {{ width: 80px; }}  /* Permite WhatsApp */
            table.tabla-membresias th:nth-child(12),
            table.tabla-membresias td:nth-child(12) {{ width: 100px; text-align: left; }}  /* Registrado por */
            
            /* ANCHOS DE COLUMNA ESPECÍFICOS PARA REPORTE DE EMPLEADOS */
            table.tabla-empleados {{ 
                width: 100%; 
                table-layout: fixed; 
            }}
            table.tabla-empleados th,
            table.tabla-empleados td {{
                padding: 8px 6px;
                font-size: 9pt;
                vertical-align: middle;
                line-height: 1.4;
                word-wrap: break-word;
                overflow-wrap: break-word;
            }}
            table.tabla-empleados th {{
                font-size: 9pt;
                font-weight: bold;
                white-space: normal;
            }}
            /* Las columnas de empleados usan porcentajes para distribución flexible */
            table.tabla-empleados th:nth-child(1),
            table.tabla-empleados td:nth-child(1) {{ text-align: center; }}  /* N° */
            table.tabla-empleados th:nth-child(2),
            table.tabla-empleados td:nth-child(2) {{ text-align: left; }}  /* Nombre */
            table.tabla-empleados th:nth-child(3),
            table.tabla-empleados td:nth-child(3) {{ text-align: center; }}  /* DNI */
            table.tabla-empleados th:nth-child(4),
            table.tabla-empleados td:nth-child(4) {{ text-align: center; }}  /* Teléfono */
            table.tabla-empleados th:nth-child(5),
            table.tabla-empleados td:nth-child(5) {{ text-align: left; }}  /* Email */
            table.tabla-empleados th:nth-child(6),
            table.tabla-empleados td:nth-child(6) {{ text-align: center; }}  /* Rol */
            table.tabla-empleados th:nth-child(7),
            table.tabla-empleados td:nth-child(7) {{ text-align: center; }}  /* Fecha Contratación */
            table.tabla-empleados th:nth-child(8),
            table.tabla-empleados td:nth-child(8) {{ text-align: center; }}  /* Estado */
            table.tabla-empleados th:nth-child(9),
            table.tabla-empleados td:nth-child(9) {{ text-align: left; }}  /* Registrado por */
            
            /* FOOTER - MANTENIDO IGUAL */
            .footer {{
                margin-top: 12px;
                padding-top: 8px;
                border-top: 1px solid {self.COLORES['borde']};
                text-align: center;
                font-size: 8pt;
                color: {self.COLORES['texto_claro']};
            }}
            
            .badge {{
                display: inline-block;
                padding: 2px 8px;
                border-radius: 10px;
                font-size: 8pt;
                font-weight: 500;
            }}
            
            /* CLASE ESPECÍFICA SOLO PARA REDUCIR ESPACIO LOGO-TEXTO */
            .logo-cerca-texto {{
                margin-bottom: 0 !important;
                padding-bottom: 0 !important;
            }}
            
            .texto-cerca-logo {{
                margin-top: 0 !important;
                padding-top: 0 !important;
            }}
            
            /* NUEVO: ESTILOS PARA TABLA DE HISTORIAL DE MEMBRESÍAS - DISEÑO DIFERENCIADO */
            .historial-section {{
                background-color: #F3F4F6;
                border-left: 4px solid #6B7280;
                margin: 15px 0;
                padding: 15px;
                page-break-inside: avoid;
            }}
            
            .historial-title {{
                font-size: 14px;
                font-weight: bold;
                color: #374151;
                margin-bottom: 10px;
                display: flex;
                align-items: center;
            }}
            
            .historial-title::before {{
                content: "📋";
                margin-right: 8px;
            }}
            
            .historial-info {{
                display: flex;
                gap: 20px;
                margin-bottom: 12px;
                font-size: 10px;
                color: #6B7280;
            }}
            
            .historial-table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
                font-size: 8pt;
                background-color: #FFFFFF;
                border: 1px solid #D1D5DB;
            }}
            
            .historial-table th {{
                background: #6B7280 !important;  /* Gris oscuro - diferente del primary */
                color: white !important;
                padding: 6px 8px;
                text-align: left;
                font-weight: bold;
                font-size: 8pt;
                border-bottom: 2px solid #4B5563;
            }}
            
            .historial-table td {{
                padding: 5px 8px;
                border-bottom: 1px solid #E5E7EB;
                font-size: 8pt;
                color: #374151;
            }}
            
            .historial-table tr:nth-child(even) {{
                background: #F9FAFB;
            }}
            
            .historial-table tr:hover {{
                background: #E5E7EB;
            }}
            
            .historial-badge {{
                display: inline-block;
                padding: 2px 6px;
                border-radius: 4px;
                font-size: 7pt;
                font-weight: 500;
            }}
            
            .historial-badge-activa {{
                background-color: #D1FAE5;
                color: #065F46;
            }}
            
            .historial-badge-pagado {{
                background-color: #D1FAE5;
                color: #065F46;
            }}
            
            .historial-badge-pendiente {{
                background-color: #FEF3C7;
                color: #92400E;
            }}
            
            .historial-badge-terminado {{
                background-color: #E5E7EB;
                color: #374151;
            }}
            
            .historial-badge-vencida {{
                background-color: #FEE2E2;
                color: #991B1B;
            }}
            
            .historial-badge-terminada {{
                background-color: #E5E7EB;
                color: #374151;
            }}
            
            .historial-badge-terminado {{
                background-color: #E5E7EB;
                color: #374151;
            }}
            
            .historial-badge-cancelada {{
                background-color: #FEF3C7;
                color: #92400E;
            }}
            
            /* Badge states */
            .badge-success {{
                background-color: #D1FAE5;
                color: #065F46;
            }}
            
            .badge-warning {{
                background-color: #FEF3C7;
                color: #92400E;
            }}
            
            .badge-info {{
                background-color: #DBEAFE;
                color: #1E40AF;
            }}
        """
    
        return html
    def generar_header_centrado(self, titulo_reporte, subtitulo=None):
        """Genera HTML para el header centrado con logo grande y diseño compacto"""
        logo_base64 = self.get_logo_base64()
        
        logo_html = ""
        if logo_base64:
            logo_html = f'''
            <div class="logo-section tight-spacing">
                <img src="{logo_base64}" alt="Logo {self.config['empresa_nombre']}" 
                    style="height: 140px; max-width: 400px;">
            </div>
            '''
        else:
            logo_html = f'''
            <div class="logo-section tight-spacing">
                <div class="logo-placeholder">
                    {self.config['empresa_nombre'][:20]}
                </div>
            </div>
            '''
        
        subtitulo_text = f"<p class='no-margin'>{subtitulo}</p>" if subtitulo else ""
        
        return f"""
        <div class="header-container compact">
            {logo_html}
            <div class="company-info tight-spacing">
                <h1 class="no-top-space">{self.config['empresa_nombre']}</h1>
                <p class="no-top-space">Sistema de Gestión Deportiva</p>
            </div>
            <div class="report-meta tight-spacing">
                <p class="no-margin"><strong>REPORTE:</strong> {titulo_reporte}</p>
                {subtitulo_text}
                <p class="no-margin"><strong>FECHA:</strong> {datetime.now().strftime('%d/%m/%Y')}</p>
                <p class="no-margin"><strong>HORA:</strong> {datetime.now().strftime('%H:%M')}</p>
            </div>
        </div>
        """
    
    
    def obtener_detalles_ventas_por_ids(self, venta_ids):
        """
        Obtiene los detalles de productos para cada venta seleccionada
        Similar a obtener_entradas_inventario_por_ids() para productos
        
        Args:
            venta_ids: Lista de IDs de ventas
        
        Returns:
            Diccionario con estructura:
            {
                '123': {
                    'detalles': [
                        {
                            'producto_nombre': 'Proteína',
                            'categoria': 'Suplementos',
                            'cantidad': 2,
                            'precio_unitario': 120.00,
                            'subtotal': 240.00
                        },
                        ...
                    ],
                    'info': {
                        'fecha_venta': '2026-02-13 14:30:00',
                        'cliente_nombre': 'Juan Pérez',
                        'metodo_pago': 'Efectivo',
                        'empleado': 'María García',
                        'total': 240.00
                    }
                },
                ...
            }
        """
        if not venta_ids or len(venta_ids) == 0:
            return {}
        
        conn = self._get_connection()
        cursor = conn.cursor()
        
        detalles_por_venta = {}
        
        for venta_id in venta_ids:
            # Obtener detalles de productos vendidos
            cursor.execute('''
                SELECT 
                    dv.producto_id,
                    dv.cantidad,
                    dv.precio_unitario,
                    dv.subtotal,
                    p.nombre as producto_nombre,
                    p.categoria
                FROM detalle_ventas dv
                JOIN productos p ON dv.producto_id = p.id
                WHERE dv.venta_id = %s
                ORDER BY dv.id
            ''', (venta_id,))
            
            detalles = [dict(row) for row in cursor.fetchall()]
            
            # Obtener información general de la venta
            cursor.execute('''
                SELECT 
                    v.id,
                    v.fecha_venta,
                    v.total,
                    v.metodo_pago,
                    v.estado,
                    v.cliente_nombre,
                    u.nombre_completo as empleado
                FROM ventas v
                LEFT JOIN usuarios u ON v.usuario_id = u.id
                WHERE v.id = %s
            ''', (venta_id,))
            
            venta_info = cursor.fetchone()
            
            if venta_info:
                detalles_por_venta[str(venta_id)] = {
                    'detalles': detalles,
                    'info': dict(venta_info),
                    'venta_id': venta_id,
                    'fecha': venta_info['fecha_venta'],
                    'cliente': venta_info['cliente_nombre'] or 'Cliente General',
                    'total': venta_info['total']
                }
        
        conn.close()
        return detalles_por_venta
    
    def generar_html_con_detalles_ventas(self, tipo_reporte, datos_tabla, detalles_ventas, landscape=False):
        """
        Genera HTML para reportes de ventas con detalles de productos vendidos
        VERSIÓN CORREGIDA: Detalles DENTRO de la tabla (no al final)
        """
        from datetime import datetime
        
        logo_base64 = self.get_logo_base64()
        headers = datos_tabla.get('headers', [])
        rows = datos_tabla.get('rows', [])
        rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
        
        # ===== HEADER =====
        header_html = f"""
        <div class="header-container">
            <div class="logo-section">
                {f'<img src="{logo_base64}" alt="Logo" style="height: 140px; max-width: 400px;">' if logo_base64 else ''}
            </div>
        </div>
        """
        
        # ===== CALCULAR ESTADÍSTICAS =====
        total_ventas = len(rows)
        total_productos_vendidos = sum(
            len(detalles_ventas[str(r.get('id'))]['detalles']) 
            for r in rows_con_detalles 
            if str(r.get('id')) in detalles_ventas
        )
        monto_total = sum(
            detalles_ventas[str(r.get('id'))]['total']
            for r in rows_con_detalles 
            if str(r.get('id')) in detalles_ventas
        )
        
        # ===== TABLA PRINCIPAL CON DETALLES INTEGRADOS =====
        tabla_ventas = f"""
        <table>
            <thead>
                <tr>
        """
        for header in headers:
            tabla_ventas += f"<th>{header}</th>"
        tabla_ventas += """
                </tr>
            </thead>
            <tbody>
        """
        
        # ✅ CORRECCIÓN: Insertar detalles DENTRO del tbody
        for fila_con_detalles in rows_con_detalles:
            row = fila_con_detalles.get('datos', [])
            venta_id = fila_con_detalles.get('id')
            
            # Fila principal de la venta
            tabla_ventas += "<tr>"
            for cell in row:
                cell_value = str(cell) if cell is not None else ''
                tabla_ventas += f"<td>{cell_value}</td>"
            tabla_ventas += "</tr>"
            
            # ✅ Insertar detalle JUSTO DESPUÉS de la fila (DENTRO de la tabla)
            if venta_id and str(venta_id) in detalles_ventas:
                venta_data = detalles_ventas[str(venta_id)]
                detalles = venta_data.get('detalles', [])
                info = venta_data.get('info', {})
                
                if detalles and len(detalles) > 0:
                    total_venta = info.get('total', 0)
                    
                    # Fila de detalle (colspan para ocupar todas las columnas)
                    tabla_ventas += f"""
                <tr>
                    <td colspan="{len(headers)}" style="padding: 0; background-color: #EFF6FF; border-left: 4px solid #3B82F6;">
                        <div style="padding: 10px 15px;">
                            <table style="width: 100%; font-size: 8pt; background-color: #FFFFFF; border: 1px solid #D1D5DB; margin: 0; border-collapse: collapse;">
                                <thead>
                                    <tr style="background: #3B82F6; color: white;">
                                        <th style="padding: 6px 8px; text-align: left; border-bottom: 2px solid #2563EB;">Producto</th>
                                        <th style="padding: 6px 8px; text-align: left; border-bottom: 2px solid #2563EB;">Categoría</th>
                                        <th style="padding: 6px 8px; text-align: center; border-bottom: 2px solid #2563EB;">Cantidad</th>
                                        <th style="padding: 6px 8px; text-align: right; border-bottom: 2px solid #2563EB;">Precio Unit.</th>
                                        <th style="padding: 6px 8px; text-align: right; border-bottom: 2px solid #2563EB;">Subtotal</th>
                                    </tr>
                                </thead>
                                <tbody>
                    """
                    
                    for idx, detalle in enumerate(detalles):
                        bg_color = '#F9FAFB' if idx % 2 == 0 else '#FFFFFF'
                        producto_nombre = detalle.get('producto_nombre', 'Producto')
                        categoria = detalle.get('categoria', 'Sin categoría')
                        cantidad = detalle.get('cantidad', 0)
                        precio_unitario = detalle.get('precio_unitario', 0)
                        subtotal = detalle.get('subtotal', 0)
                        
                        tabla_ventas += f"""
                                    <tr style="background: {bg_color}; border-bottom: 1px solid #E5E7EB;">
                                        <td style="padding: 5px 8px;">{producto_nombre}</td>
                                        <td style="padding: 5px 8px;">{categoria}</td>
                                        <td style="padding: 5px 8px; text-align: center;">
                                            <span style="background-color: #DBEAFE; color: #1E40AF; padding: 2px 6px; border-radius: 4px; font-weight: bold;">
                                                x{cantidad}
                                            </span>
                                        </td>
                                        <td style="padding: 5px 8px; text-align: right;">S/. {precio_unitario:.2f}</td>
                                        <td style="padding: 5px 8px; text-align: right; font-weight: bold;">S/. {subtotal:.2f}</td>
                                    </tr>
                        """
                    
                    # Fila de total de la venta
                    tabla_ventas += f"""
                                    <tr style="background: #EFF6FF; border-top: 2px solid #3B82F6;">
                                        <td colspan="4" style="padding: 8px; text-align: right; font-weight: bold; color: #1E40AF;">TOTAL VENTA:</td>
                                        <td style="padding: 8px; text-align: right; font-weight: bold; color: #1E40AF;">S/. {total_venta:.2f}</td>
                                    </tr>
                                </tbody>
                            </table>
                        </div>
                    </td>
                </tr>
                    """
        
        tabla_ventas += """
            </tbody>
        </table>
        """
        
        # ===== CSS =====
        css = self.generar_css_pdf(landscape)
        
        # ===== HTML COMPLETO =====
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{css}</style>
        </head>
        <body>
            {header_html}
            
            <div class="content">
                <h2 class="report-title">Reporte de Ventas</h2>
                
                <table class="stats-table" cellpadding="0" cellspacing="0">
                    <tr>
                        <td class="stat-cell">
                            <div class="stat-label">Total Ventas</div>
                            <div class="stat-value">{total_ventas}</div>
                        </td>
                        <td class="stat-cell stat-cell-middle">
                            <div class="stat-label">Productos Vendidos</div>
                            <div class="stat-value">{total_productos_vendidos}</div>
                        </td>
                        <td class="stat-cell">
                            <div class="stat-label">Monto Total</div>
                            <div class="stat-value">S/. {monto_total:.2f}</div>
                        </td>
                    </tr>
                </table>
                
                {tabla_ventas}
                
                <div class="footer">
                    <p><strong>{self.config['empresa_nombre']}</strong> - Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html


    
    
    
    
    def generar_html_desde_tabla(self, tipo_reporte, headers, rows, details=None, landscape=False):
        """
        Genera HTML desde los datos de una tabla (headers y rows)
        Args:
            tipo_reporte: Tipo de reporte (clientes, ventas, productos, etc.)
            headers: Lista de cabeceras de la tabla
            rows: Lista de filas de datos
            details: Lista opcional de diccionarios con detalles para cada fila
            landscape: Si True, usa orientación horizontal
        """
        logo_base64 = self.get_logo_base64()
        
        # Obtener nombre de empresa para el título
        empresa_nombre = self.config.get('empresa_nombre', 'Empresa')
        
        # Definir anchos de columna para reportes de membresías - USANDO ANCHO COMPLETO
        es_membresia = tipo_reporte == "membresias"
        es_empleados = tipo_reporte == "empleados"

        # Anchos optimizados para cada tipo de reporte
        anchos_membresia = ['30', '60', '150', '65', '65', '90', '55', '110', '110', '100', '100', '130'] if es_membresia else []
        # Anchos mejorados para empleados en porcentajes: N°(5%), Nombre(18%), DNI(10%), Teléfono(12%), Email(20%), Rol(10%), Fecha(12%), Estado(8%), Registrado(15%)
        anchos_empleados = ['5%', '18%', '10%', '12%', '20%', '15%', '12%', '8%', '15%'] if es_empleados else []
        ancho_tabla = '100%'
        
        # Generar tabla HTML
        table_html = ""
        
        # Headers de la tabla
        table_html += f'<thead><tr>'
        for idx, header in enumerate(headers):
            if es_membresia and idx < len(anchos_membresia):
                ancho = anchos_membresia[idx]
                table_html += f'<th style="width: {ancho}px;">{header}</th>'
            else:
                table_html += f"<th>{header}</th>"
        table_html += '</tr></thead><tbody>'
        
        # Filas principales
        for idx, row in enumerate(rows):
            table_html += "<tr>"
            for col_idx, cell in enumerate(row):
                if es_membresia and col_idx < len(anchos_membresia):
                    ancho = anchos_membresia[col_idx]
                    align = "left" if col_idx in [2, 11] else "center"
                    table_html += f'<td style="width: {ancho}px; text-align: {align};">{cell if cell is not None else ""}</td>'
                elif es_empleados and col_idx < len(anchos_empleados):
                    ancho = anchos_empleados[col_idx]
                    # Alineaciones específicas para empleados
                    if col_idx == 0:  # N°
                        align = "center"
                    elif col_idx == 4:  # Email
                        align = "left"
                    elif col_idx == 7:  # Estado
                        align = "center"
                    else:
                        align = "left"
                    table_html += f'<td style="width: {ancho}px; text-align: {align}; word-wrap: break-word;">{cell if cell is not None else ""}</td>'
            table_html += "</tr>"
            
            # Si hay detalles para esta fila, agregarlos
            if details and idx < len(details) and details[idx]:
                detail = details[idx]
                table_html += self._generar_html_detalles(tipo_reporte, detail)
        
        table_html += """
                    </tbody>
        """
        
        # Generar clase específica para el tipo de reporte
        if tipo_reporte == "membresias":
            tabla_clase = "tabla-membresias"
        elif tipo_reporte == "empleados":
            tabla_clase = "tabla-empleados"
        else:
            tabla_clase = ""
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{self.generar_css_pdf(landscape)}</style>
        </head>
        <body>
            <div class="header-container">
                <div class="logo-section">
                    {f'<img src="{logo_base64}" alt="Logo">' if logo_base64 else ''}
                </div>
            </div>
            
            <div class="content">
                <h2 class="report-title">Reporte de {tipo_reporte.capitalize()}</h2>
                
                <table class="stats-table" cellpadding="0" cellspacing="0">
                    <tr>
                        <td class="stat-cell stat-cell-wide">
                            <div class="stat-label">Total Registros</div>
                            <div class="stat-value">{len(rows)}</div>
                        </td>
                        <td class="stat-cell stat-cell-wide">
                            <div class="stat-label">Fecha Generación</div>
                            <div class="stat-value">{datetime.now().strftime('%d/%m/%Y')}</div>
                        </td>
                    </tr>
                </table>
                
                <table class="{tabla_clase}" style="width: 100%; table-layout: fixed; margin: 0;">

                    {table_html}
                </table>
                
                <div class="footer">
                    <p>{empresa_nombre} - Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generar_html_detalles(self, tipo_reporte, detail):
        """Genera el HTML para los detalles de una fila"""
        if not detail:
            return ""
        
        detail_html = ""
        
        # Determinar el tipo de detalle según el tipo de reporte
        if tipo_reporte == 'clientes' and detail.get('historial'):
            # Detalles de historial de cliente
            historial = detail.get('historial', [])
            accesos = detail.get('accesos', [])
            
            detail_html = f"""
                <tr class="detail-section no-break">
                    <td colspan="{len(detail.get('row_data', [])) if detail.get('row_data') else 9}">
                        <div class="detail-header">Historial de Pagos y Membresías</div>
                        <div class="detail-info">
                            <div class="detail-info-item"><strong>Total Pagos:</strong> {len(historial)}</div>
                            <div class="detail-info-item"><strong>Total Accesos:</strong> {len(accesos)}</div>
                        </div>
                        <table class="detail-table">
                            <thead>
                                <tr>
                                    <th>Fecha Pago</th>
                                    <th>Plan</th>
                                    <th>Monto</th>
                                    <th>Método</th>
                                    <th>Estado</th>
                                    <th>Registro</th>
                                </tr>
                            </thead>
                            <tbody>
            """
            
            for pago in historial[:15]:  # Limitar a 15 registros
                estado = pago.get('estado', '')
                badge_class = 'badge-success' if estado == 'completado' else ('badge-warning' if estado == 'pendiente' else 'badge-info')
                detail_html += f"""
                                <tr>
                                    <td>{pago.get('fecha_pago', '')}</td>
                                    <td>{pago.get('plan_nombre', '')}</td>
                                    <td>S/. {float(pago.get('monto_pagado', 0)):.2f}</td>
                                    <td>{pago.get('metodo_pago', '')}</td>
                                    <td><span class="badge {badge_class}">{estado}</span></td>
                                    <td>{pago.get('fecha_registro', '')}</td>
                                </tr>
                """
            
            if accesos:
                detail_html += """
                            </tbody>
                        </table>
                        <div class="detail-header" style="margin-top: 15px;">Últimos Accesos</div>
                        <table class="detail-table">
                            <thead>
                                <tr>
                                    <th>Fecha/Hora</th>
                                    <th>Tipo</th>
                                    <th>Estado</th>
                                </tr>
                            </thead>
                            <tbody>
                """
                for acceso in accesos[:5]:  # Últimos 5 accesos
                    detail_html += f"""
                                <tr>
                                    <td>{acceso.get('fecha_hora_entrada', '')}</td>
                                    <td>{acceso.get('tipo_acceso', 'Regular')}</td>
                                    <td>{acceso.get('estado', 'Completado')}</td>
                                </tr>
                    """
            
            detail_html += """
                            </tbody>
                        </table>
                    </td>
                </tr>
            """
        
        elif tipo_reporte == 'ventas' and detail.get('detalles'):
            # Detalles de venta
            detalles = detail.get('detalles', [])
            venta_info = detail.get('info', {})
            
            detail_html = f"""
                <tr class="detail-section no-break">
                    <td colspan="{len(detail.get('row_data', [])) if detail.get('row_data') else 9}">
                        <div class="detail-header">Detalle de Venta #{venta_info.get('id', '')}</div>
                        <div class="detail-info">
                            <div class="detail-info-item"><strong>Fecha:</strong> {venta_info.get('fecha', '')}</div>
                            <div class="detail-info-item"><strong>Cliente:</strong> {venta_info.get('cliente', 'Consumidor final')}</div>
                            <div class="detail-info-item"><strong>Método Pago:</strong> {venta_info.get('metodo_pago', 'Efectivo')}</div>
                            <div class="detail-info-item"><strong>Empleado:</strong> {venta_info.get('empleado', 'Sistema')}</div>
                        </div>
                        <table class="detail-table">
                            <thead>
                                <tr>
                                    <th>Producto</th>
                                    <th>Categoría</th>
                                    <th>Cantidad</th>
                                    <th>Precio Unit.</th>
                                    <th>Subtotal</th>
                                </tr>
                            </thead>
                            <tbody>
            """
            
            for prod in detalles:
                detail_html += f"""
                                <tr>
                                    <td>{prod.get('nombre', '')}</td>
                                    <td>{prod.get('categoria', '')}</td>
                                    <td>{prod.get('cantidad', 0)}</td>
                                    <td>S/. {float(prod.get('precio_unitario', 0)):.2f}</td>
                                    <td>S/. {float(prod.get('subtotal', 0)):.2f}</td>
                                </tr>
                """
            
            total = venta_info.get('total', 0)
            detail_html += f"""
                            </tbody>
                            <tfoot>
                                <tr>
                                    <td colspan="4" style="text-align: right; font-weight: bold;">TOTAL:</td>
                                    <td style="font-weight: bold; font-size: 10pt;">S/. {total:.2f}</td>
                                </tr>
                            </tfoot>
                        </table>
                    </td>
                </tr>
            """
        
        elif tipo_reporte == 'productos' and detail.get('entradas'):
            # Detalles de entradas de inventario
            entradas = detail.get('entradas', [])
            
            detail_html = f"""
                <tr class="detail-section no-break">
                    <td colspan="{len(detail.get('row_data', [])) if detail.get('row_data') else 9}">
                        <div class="detail-header">Historial de Entradas de Inventario</div>
                        <div class="detail-info">
                            <div class="detail-info-item"><strong>Total Entradas:</strong> {len(entradas)}</div>
                            <div class="detail-info-item"><strong>Total Unidades:</strong> {sum(e.get('cantidad', 0) for e in entradas)}</div>
                        </div>
                        <table class="detail-table">
                            <thead>
                                <tr>
                                    <th>Fecha Entrada</th>
                                    <th>Cantidad</th>
                                    <th>Costo Unitario</th>
                                    <th>Subtotal</th>
                                    <th>Registrado por</th>
                                </tr>
                            </thead>
                            <tbody>
            """
            
            for entrada in reversed(list(entradas[:15])):
                subtotal = float(entrada.get('cantidad', 0) or 0) * float(entrada.get('costo_unitario', 0) or 0)
                detail_html += f"""
                                <tr>
                                    <td>{entrada.get('fecha_entrada', '')}</td>
                                    <td>+{entrada.get('cantidad', 0)}</td>
                                    <td>S/. {float(entrada.get('costo_unitario', 0)):.2f}</td>
                                    <td>S/. {subtotal:.2f}</td>
                                    <td>{entrada.get('usuario_registro', 'Sistema')}</td>
                                </tr>
                """
            
            detail_html += """
                            </tbody>
                        </table>
                    </td>
                </tr>
            """
        
        return detail_html
    
    def generar_html_con_historial(self, tipo_reporte, datos_tabla, historial_membresias, landscape=False):
        """
        Genera HTML para reportes con historial de membresías incluido
        Diseño diferenciado para la tabla de historial
        
        Args:
            tipo_reporte: Tipo de reporte (membresias, clientes, etc.)
            datos_tabla: Diccionario con headers, rows y rows_con_detalles
            historial_membresias: Diccionario con el historial de membresías por cliente_id
            landscape: Si True, usa orientación horizontal
        """
        logo_base64 = self.get_logo_base64()
        
        headers = datos_tabla.get('headers', [])
        rows = datos_tabla.get('rows', [])
        rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
        
        # Obtener nombre de empresa para el título
        empresa_nombre = self.config.get('empresa_nombre', 'Empresa')
        
        # Definir anchos de columna para reportes de membresías - USANDO ANCHO COMPLETO
        es_membresia = tipo_reporte == "membresias"
        es_empleados = tipo_reporte == "empleados"

        # Anchos optimizados para cada tipo de reporte
        anchos_membresia = ['30', '60', '150', '65', '65', '90', '55', '110', '110', '100', '100', '130'] if es_membresia else []
        # Anchos mejorados para empleados en porcentajes: N°(5%), Nombre(18%), DNI(10%), Teléfono(12%), Email(20%), Rol(10%), Fecha(12%), Estado(8%), Registrado(15%)
        anchos_empleados = ['5%', '18%', '10%', '12%', '20%', '15%', '12%', '8%', '15%'] if es_empleados else []
        ancho_tabla = '100%'
        
        # Generar tabla HTML principal
        table_html = ""
        
        # Headers de la tabla principal
        table_html += '<thead><tr>'
        for idx, header in enumerate(headers):
            if es_membresia and idx < len(anchos_membresia):
                ancho = anchos_membresia[idx]
                table_html += f'<th style="width: {ancho}px;">{header}</th>'
            elif es_empleados and idx < len(anchos_empleados):
                ancho = anchos_empleados[idx]
                table_html += f'<th style="width: {ancho}px;">{header}</th>'
            else:
                table_html += f"<th>{header}</th>"
        table_html += '</tr></thead><tbody>'
        clientes_con_historial = sum(
            1 for r in rows_con_detalles 
            if str(r.get('id')) in historial_membresias 
            and len(historial_membresias[str(r.get('id'))].get('historial', [])) > 1
        )
        # Filas principales con sus historiales asociados
        for idx, fila_con_detalles in enumerate(rows_con_detalles):
            row = fila_con_detalles.get('datos', [])
            cliente_id = fila_con_detalles.get('id')
            
            # Fila principal
            table_html += "<tr>"
            for col_idx, cell in enumerate(row):
                if es_membresia and col_idx < len(anchos_membresia):
                    ancho = anchos_membresia[col_idx]
                    align = "left" if col_idx in [2, 11] else "center"
                    table_html += f'<td style="width: {ancho}px; text-align: {align};">{cell if cell is not None else ""}</td>'
                elif es_empleados and col_idx < len(anchos_empleados):
                    ancho = anchos_empleados[col_idx]
                    # Alineaciones específicas para empleados
                    if col_idx == 0:  # N°
                        align = "center"
                    elif col_idx == 4:  # Email
                        align = "left"
                    elif col_idx == 7:  # Estado
                        align = "center"
                    else:
                        align = "left"
                    table_html += f'<td style="width: {ancho}px; text-align: {align}; word-wrap: break-word;">{cell if cell is not None else ""}</td>'
                else:
                    table_html += f"<td>{cell if cell is not None else ''}</td>"
            table_html += "</tr>"
            
            # Si hay historial de membresías para este cliente, agregarlo
            if historial_membresias and str(cliente_id) in historial_membresias:
                historial_data = historial_membresias[str(cliente_id)]
                historial = historial_data.get('historial', [])
                cliente_nombre = historial_data.get('cliente_nombre', '')
                cliente_dni = historial_data.get('cliente_dni', '')
                
                if historial and len(historial) > 0:
                    # Generar HTML para la tabla de historial con diseño diferenciado
                    historial_html = self._generar_html_tabla_historial(
                        cliente_id=cliente_id,
                        cliente_nombre=cliente_nombre,
                        cliente_dni=cliente_dni,
                        historial=historial,
                        colspan=len(headers)
                    )
                    table_html += historial_html
        
        table_html += """
                    </tbody>
        """
        
        # Generar clase específica para el tipo de reporte
        if tipo_reporte == "membresias":
            tabla_clase = "tabla-membresias"
        elif tipo_reporte == "empleados":
            tabla_clase = "tabla-empleados"
        else:
            tabla_clase = ""
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{self.generar_css_pdf(landscape)}</style>
        </head>
        <body>
            <div class="header-container">
                <div class="logo-section">
                    {f'<img src="{logo_base64}" alt="Logo">' if logo_base64 else ''}
            
                </div>
            </div>
            
            <div class="content">
                <h2 class="report-title">Reporte de {tipo_reporte.capitalize()}</h2>
                
                <table class="stats-table" cellpadding="0" cellspacing="0">
                    <tr>
                        <td class="stat-cell">
                            <div class="stat-label">Total Registros</div>
                            <div class="stat-value">{len(rows)}</div>
                        </td>
                        <td class="stat-cell stat-cell-middle">
                            <div class="stat-label">Renovaciones</div>
                            <div class="stat-value">{clientes_con_historial}</div>
                        </td>
                        <td class="stat-cell">
                            <div class="stat-label">Fecha Generación</div>
                            <div class="stat-value">{datetime.now().strftime('%d/%m/%Y')}</div>
                        </td>
                    </tr>
                </table>
                
                <table class="{tabla_clase}">
                    {table_html}
                </table>
                
                <div class="footer">
                    <p>{empresa_nombre} - Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def generar_html_con_entradas(self, tipo_reporte, datos_tabla, entradas_inventario, landscape=False):
        """
        Genera HTML para reportes de productos con entradas de inventario
        VERSIÓN MEJORADA: Mejor diseño visual y espaciado
        """
        from datetime import datetime
        
        logo_base64 = self.get_logo_base64()
        headers = datos_tabla.get('headers', [])
        rows = datos_tabla.get('rows', [])
        rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
        
        # ===== HEADER =====
        header_html = f"""
        <div class="header-container">
            <div class="logo-section">
                {f'<img src="{logo_base64}" alt="Logo" style="height: 140px; max-width: 400px;">' if logo_base64 else ''}
            </div>
        </div>
        """
        
        # ===== CALCULAR ESTADÍSTICAS =====
        total_entradas = sum(
            len(entradas_inventario[str(r.get('id'))]) 
            for r in rows_con_detalles 
            if str(r.get('id')) in entradas_inventario
        )
        
        # ===== TABLA PRINCIPAL CON ENTRADAS INTEGRADAS =====
        tabla_productos = f"""
        <table>
            <thead>
                <tr>
        """
        for header in headers:
            tabla_productos += f"<th>{header}</th>"
        tabla_productos += """
                </tr>
            </thead>
            <tbody>
        """
        
        # ✅ CORRECCIÓN: Insertar entradas DENTRO del tbody
        for fila_con_detalles in rows_con_detalles:
            row = fila_con_detalles.get('datos', [])
            producto_id = fila_con_detalles.get('id')
            
            # Fila principal del producto
            tabla_productos += "<tr>"
            for cell in row:
                cell_value = str(cell) if cell is not None else ''
                tabla_productos += f"<td>{cell_value}</td>"
            tabla_productos += "</tr>"
            
            # ✅ Insertar entradas JUSTO DESPUÉS de la fila (DENTRO de la tabla)
            if producto_id and str(producto_id) in entradas_inventario:
                entradas = entradas_inventario[str(producto_id)]
                if entradas and len(entradas) > 0:
                    # Buscar nombre del producto
                    nombre_producto = row[2] if len(row) > 2 else f"Producto #{producto_id}"
                    nombre = row[1] if len(row) > 1 else f"Producto #{producto_id}"
                    
                    # 🎨 DISEÑO MEJORADO
                    tabla_productos += f"""
                <tr>
                    <td colspan="{len(headers)}" style="padding: 0; background-color: #F8FAFC; border-left: 4px solid #10B981;">
                        <div style="padding: 15px 20px;">
                            
                            <!-- TÍTULO CON MEJOR ESPACIADO -->
                            <div style="
                                font-size: 12px; 
                                font-weight: bold; 
                                color: #047857; 
                                margin-bottom: 12px; 
                                padding: 8px 12px;
                                background-color: #D1FAE5;
                                border-radius: 6px;
                                display: inline-block;
                            ">
                                📦 Historial de Entradas: {nombre_producto}
                            </div>
                            
                            
                            <!-- TABLA CON DISEÑO MEJORADO -->
                            <table style="
                                width: 100%; 
                                font-size: 8pt; 
                                background-color: #FFFFFF; 
                                border: 1px solid #CBD5E1; 
                                margin: 0; 
                                border-collapse: collapse;
                                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                            ">
                                <thead>
                                    <tr style="background: #10B981; color: white;">
                                        <th style="padding: 8px 10px; text-align: left; border-bottom: 2px solid #059669; font-size: 9pt;">Fecha</th>
                                        <th style="padding: 8px 10px; text-align: center; border-bottom: 2px solid #059669; font-size: 9pt;">Cantidad</th>
                                        <th style="padding: 8px 10px; text-align: right; border-bottom: 2px solid #059669; font-size: 9pt;">Costo Unit.</th>
                                        <th style="padding: 8px 10px; text-align: right; border-bottom: 2px solid #059669; font-size: 9pt;">Subtotal</th>
                                        <th style="padding: 8px 10px; text-align: left; border-bottom: 2px solid #059669; font-size: 9pt;">Registrado por</th>
                                    </tr>
                                </thead>
                                <tbody>
                    """
                    
                    for idx, entrada in enumerate(reversed(list(entradas[:15]))):
                        fecha = entrada.get('fecha_entrada', '')
                        if fecha and ' ' in fecha:
                            fecha = fecha.split(' ')[0]
                        
                        bg_color = '#FFFFFF' if idx % 2 == 0 else '#F8FAFC'
                        
                        tabla_productos += f"""
                                    <tr style="background: {bg_color}; border-bottom: 1px solid #E2E8F0;">
                                        <td style="padding: 7px 10px; font-size: 8pt;">{fecha}</td>
                                        <td style="padding: 7px 10px; text-align: center;">
                                            <span style="
                                                background-color: #D1FAE5; 
                                                color: #065F46; 
                                                padding: 3px 8px; 
                                                border-radius: 4px; 
                                                font-weight: bold;
                                                font-size: 8pt;
                                            ">
                                                +{entrada.get('cantidad', 0)}
                                            </span>
                                        </td>
                                        <td style="padding: 7px 10px; text-align: right; font-size: 8pt;">S/. {float(entrada.get('costo_unitario', 0)):.2f}</td>
                                        <td style="padding: 7px 10px; text-align: right; font-weight: bold; font-size: 8pt;">S/. {float(entrada.get('cantidad', 0) or 0) * float(entrada.get('costo_unitario', 0) or 0):.2f}</td>
                                        <td style="padding: 7px 10px; font-size: 8pt; color: #475569;">{entrada.get('usuario_registro', 'Sistema')}</td>
                                    </tr>
                        """
                    
                    if len(entradas) > 15:
                        tabla_productos += f"""
                                    <tr>
                                        <td colspan="5" style="
                                            padding: 10px 12px; 
                                            text-align: center; 
                                            background: #F1F5F9; 
                                            color: #64748B; 
                                            font-style: italic;
                                            font-size: 8pt;
                                            border-top: 2px solid #CBD5E1;
                                        ">
                                            Mostrando 15 de {len(entradas)} entradas...
                                        </td>
                                    </tr>
                        """
                    
                    tabla_productos += """
                                </tbody>
                            </table>
                        </div>
                    </td>
                </tr>
                    """
        
        tabla_productos += """
            </tbody>
        </table>
        """
        
        # ===== CSS =====
        css = self.generar_css_pdf(landscape)
        
        # ===== HTML COMPLETO =====
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{css}</style>
        </head>
        <body>
            {header_html}
            
            <div class="content">
                <h2 class="report-title">Reporte de Inventario</h2>
                
                <table class="stats-table" cellpadding="0" cellspacing="0">
                    <tr>
                        <td class="stat-cell">
                            <div class="stat-label">Total Productos</div>
                            <div class="stat-value">{len(rows)}</div>
                        </td>
                        <td class="stat-cell stat-cell-middle">
                            <div class="stat-label">Total Entradas</div>
                            <div class="stat-value">{total_entradas}</div>
                        </td>
                        <td class="stat-cell">
                            <div class="stat-label">Fecha Generación</div>
                            <div class="stat-value">{datetime.now().strftime('%d/%m/%Y')}</div>
                        </td>
                    </tr>
                </table>
                
                {tabla_productos}
                
                <div class="footer">
                    <p><strong>{self.config['empresa_nombre']}</strong> - Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html

    def _generar_html_tabla_historial(self, cliente_id, cliente_nombre, cliente_dni, historial, colspan):
        """Genera el HTML para la tabla de historial de membresías con diseño diferenciado"""
        
        # Determinar el estilo del badge según el estado
        def get_badge_class(estado):
            estado_lower = estado.lower() if estado else ''
            # NUEVOS ESTADOS: Pagado, Pendiente, Terminado, Vencido
            if 'pagado' in estado_lower:
                return 'historial-badge-pagado'
            elif 'pendiente' in estado_lower:
                return 'historial-badge-pendiente'
            elif 'activa' in estado_lower:
                return 'historial-badge-activa'
            elif 'vencida' in estado_lower or 'vencido' in estado_lower:
                return 'historial-badge-vencida'
            elif 'terminada' in estado_lower or 'terminado' in estado_lower:
                return 'historial-badge-terminado'
            elif 'cancelada' in estado_lower:
                return 'historial-badge-cancelada'
            else:
                return 'historial-badge-pendiente'
        
        # Generar filas de la tabla de historial (más reciente primero = DESC)
        historial_rows_html = ""
        for h in reversed(historial):
            estado = h.get('estado', '')
            badge_class = get_badge_class(estado)
            
            historial_rows_html += f"""
                            <tr>
                                <td>{h.get('plan_nombre', '')}</td>
                                <td>{h.get('plan_codigo', '')}</td>
                                <td>{h.get('fecha_inicio', '')}</td>
                                <td>{h.get('fecha_fin', '')}</td>
                                <td>S/. {float(h.get('monto_pagado', 0)):.2f}</td>
                                <td>{h.get('metodo_pago', '')}</td>
                                <td><span class="historial-badge {badge_class}">{estado}</span></td>
                            </tr>
            """
        
        # Generar HTML completo de la sección de historial
        historial_html = f"""
                    <tr class="no-break">
                        <td colspan="{colspan}" style="padding: 0;">
                            <div class="historial-section">
                                <div class="historial-title">Historial de Membresías</div>
                                <div class="historial-info">
                                    <span><strong>Cliente:</strong> {cliente_nombre}</span>
                                    <span><strong>DNI:</strong> {cliente_dni}</span>
                                    <span><strong>Total Registros:</strong> {len(historial)}</span>
                                </div>
                                <table class="historial-table">
                                    <thead>
                                        <tr>
                                            <th>Plan</th>
                                            <th>Código</th>
                                            <th>Fecha Inicio</th>
                                            <th>Fecha Fin</th>
                                            <th>Monto</th>
                                            <th>Método</th>
                                            <th>Estado</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {historial_rows_html}
                                    </tbody>
                                </table>
                            </div>
                        </td>
                    </tr>
        """
        
        return historial_html
    
    def generar_html_con_detalles(self, tipo_reporte, datos_tabla, landscape=False, mostrar_detalles=True):
        """
        Genera HTML para reportes con estructura jerárquica (filas con detalles)
        Args:
            tipo_reporte: Tipo de reporte (clientes, ventas, productos, etc.)
            datos_tabla: Diccionario con headers, rows y rows_con_detalles
            landscape: Si True, usa orientación horizontal
            mostrar_detalles: Si True, muestra los detalles expandidos. Si False, solo muestra la tabla principal
        """
        logo_base64 = self.get_logo_base64()
        
        headers = datos_tabla.get('headers', [])
        rows = datos_tabla.get('rows', [])
        rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
        
        # Obtener nombre de empresa para el título
        empresa_nombre = self.config.get('empresa_nombre', 'Empresa')
        
        # Definir anchos de columna para reportes de membresías - USANDO ANCHO COMPLETO
        es_membresia = tipo_reporte == "membresias"
        es_empleados = tipo_reporte == "empleados"

        # Anchos optimizados para cada tipo de reporte
        anchos_membresia = ['30', '60', '150', '65', '65', '90', '55', '110', '110', '100', '100', '130'] if es_membresia else []
        # Anchos mejorados para empleados en porcentajes: N°(5%), Nombre(18%), DNI(10%), Teléfono(12%), Email(20%), Rol(10%), Fecha(12%), Estado(8%), Registrado(15%)
        anchos_empleados = ['5%', '18%', '10%', '12%', '20%', '15%', '12%', '8%', '15%'] if es_empleados else []
        ancho_tabla = '100%'
        
        # Generar tabla HTML
        table_html = ""
        
        # Headers de la tabla
        table_html += '<thead><tr>'
        for idx, header in enumerate(headers):
            if es_membresia and idx < len(anchos_membresia):
                ancho = anchos_membresia[idx]
                table_html += f'<th style="width: {ancho}px;">{header}</th>'
            elif es_empleados and idx < len(anchos_empleados):
                ancho = anchos_empleados[idx]
                table_html += f'<th style="width: {ancho}px;">{header}</th>'
            else:
                table_html += f"<th>{header}</th>"
        table_html += '</tr></thead><tbody>'
        
        # Filas principales con sus detalles asociados
        for idx, fila_con_detalles in enumerate(rows_con_detalles):
            row = fila_con_detalles.get('datos', [])
            detalles = fila_con_detalles.get('detalles', [])
            
            # Fila principal
            table_html += "<tr>"
            for col_idx, cell in enumerate(row):
                if es_membresia and col_idx < len(anchos_membresia):
                    ancho = anchos_membresia[col_idx]
                    align = "left" if col_idx in [2, 11] else "center"
                    table_html += f'<td style="width: {ancho}px; text-align: {align};">{cell if cell is not None else ""}</td>'
                elif es_empleados and col_idx < len(anchos_empleados):
                    ancho = anchos_empleados[col_idx]
                    # Alineaciones específicas para empleados
                    if col_idx == 0:  # N°
                        align = "center"
                    elif col_idx == 4:  # Email
                        align = "left"
                    elif col_idx == 7:  # Estado
                        align = "center"
                    else:
                        align = "left"
                    table_html += f'<td style="width: {ancho}px; text-align: {align}; word-wrap: break-word;">{cell if cell is not None else ""}</td>'
                else:
                    table_html += f"<td>{cell if cell is not None else ''}</td>"
            table_html += "</tr>"
            
            # Si hay detalles para esta fila Y se solicita mostrarlos, agregarlos como filas de detalle expandidas
            if mostrar_detalles and detalles and len(detalles) > 0:
                for detalle in detalles:
                    detail_html = self._generar_html_detalles_fila(tipo_reporte, detalle, len(headers))
                    table_html += detail_html
        
        table_html += """
                    </tbody>
        """
        
        # Generar clase específica para el tipo de reporte
        if tipo_reporte == "membresias":
            tabla_clase = "tabla-membresias"
        elif tipo_reporte == "empleados":
            tabla_clase = "tabla-empleados"
        else:
            tabla_clase = ""
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{self.generar_css_pdf(landscape)}</style>
        </head>
        <body>
            <div class="header-container">
                <div class="logo-section">
                    {f'<img src="{logo_base64}" alt="Logo">' if logo_base64 else ''}
            
                </div>
            </div>
            
            <div class="content">
                <h2 class="report-title">Reporte de {tipo_reporte.capitalize()}</h2>
                
                <table class="stats-table" cellpadding="0" cellspacing="0">
                    <tr>
                        <td class="stat-cell">
                            <div class="stat-label">Total Registros</div>
                            <div class="stat-value">{len(rows)}</div>
                        </td>
                        <td class="stat-cell stat-cell-middle">
                            <div class="stat-label">Registros con Detalles</div>
                            <div class="stat-value">{sum(1 for r in rows_con_detalles if r.get('detalles') and len(r.get('detalles', [])) > 0)}</div>
                        </td>
                        <td class="stat-cell">
                            <div class="stat-label">Fecha Generación</div>
                            <div class="stat-value">{datetime.now().strftime('%d/%m/%Y')}</div>
                        </td>
                    </tr>
                </table>
                
                <table class="{tabla_clase}" style="width: 100%; table-layout: fixed; margin: 0;">
                    {table_html}
                </table>
                
                <div class="footer">
                    <p>{empresa_nombre} - Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generar_html_detalles_fila(self, tipo_reporte, detalle, colspan=9):
        """Genera el HTML para los detalles de una fila específica"""
        if not detalle:
            return ""
        
        detalle_datos = detalle.get('datos', [])
        
        detail_html = ""
        
        # Procesar diferentes tipos de estructuras de detalle
        for datos_detalle in detalle_datos:
            if isinstance(datos_detalle, dict):
                # Tiene estructura con headers y rows (tabla interna)
                titulo = datos_detalle.get('titulo', 'Detalle')
                detail_headers = datos_detalle.get('headers', [])
                detail_rows = datos_detalle.get('rows', [])
                
                detail_html += f"""
                    <tr class="detail-section no-break" style="background-color: #F9FAFB;">
                        <td colspan="{colspan}" style="padding: 10px 20px;">
                            <div class="detail-header">{titulo}</div>
                """
                
                if detail_headers and detail_rows:
                    detail_html += """
                        <table class="detail-table" style="margin-top: 8px;">
                            <thead>
                                <tr>
                    """
                    for h in detail_headers:
                        detail_html += f"<th>{h}</th>"
                    detail_html += """
                                </tr>
                            </thead>
                            <tbody>
                    """
                    
                    for row in detail_rows:
                        detail_html += "<tr>"
                        for cell in row:
                            detail_html += f"<td>{cell if cell is not None else ''}</td>"
                        detail_html += "</tr>"
                    
                    detail_html += """
                            </tbody>
                        </table>
                    """
                elif datos_detalle.get('contenido'):
                    # Solo texto plano
                    detail_html += f"<p style='margin-top: 8px; font-size: 8pt;'>{datos_detalle['contenido']}</p>"
                
                detail_html += """
                        </td>
                    </tr>
                """
            
            elif isinstance(datos_detalle, list):
                # Es una lista de valores simples
                titulo = datos_detalle[0].get('titulo', 'Información') if datos_detalle and isinstance(datos_detalle[0], dict) else 'Detalle'
                
                detail_html += f"""
                    <tr class="detail-section no-break" style="background-color: #F9FAFB;">
                        <td colspan="{colspan}" style="padding: 10px 20px;">
                            <div class="detail-header">{titulo}</div>
                """
                
                if datos_detalle and isinstance(datos_detalle[0], dict) and datos_detalle[0].get('fila'):
                    # Tiene estructura de fila con título
                    for item in datos_detalle:
                        if item.get('fila'):
                            detail_html += f"<p style='font-size: 8pt; margin: 2px 0;'>• {' | '.join(str(c) for c in item['fila'])}</p>"
                else:
                    # Lista simple
                    for item in datos_detalle:
                        if isinstance(item, str):
                            detail_html += f"<p style='font-size: 8pt; margin: 2px 0;'>{item}</p>"
                
                detail_html += """
                        </td>
                    </tr>
                """
        
        return detail_html
    
    def generar_excel_con_detalles(self, datos_tabla, filename="reporte.xlsx", sheet_name="Reporte", mostrar_detalles=True):
        """
        Genera un Excel profesional con estructura de datos jerárquica
        Args:
            datos_tabla: Diccionario con headers, rows y rows_con_detalles
            filename: Nombre del archivo
            sheet_name: Nombre de la hoja
            mostrar_detalles: Si True, incluye los detalles expandidos. Si False, solo la tabla principal
        """
        if not PANDAS_DISPONIBLE:
            raise Exception("Pandas/OpenPyXL no está disponible. Por favor instale: pip install pandas openpyxl")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]
            
            # Convertir color hexadecimal a formato Excel (sin #)
            color_primario = self.config['color_primario'].replace('#', '')
            color_borde = self.COLORES['borde'].replace('#', '')
            
            # Estilos
            header_fill = PatternFill(start_color=color_primario, 
                                       end_color=color_primario, 
                                       fill_type='solid')
            header_font = Font(color="FFFFFF", bold=True, size=11)
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(bold=True, size=11)
            data_font = Font(size=10)
            detail_font = Font(size=9, italic=True)
            border = Border(
                left=Side(style='thin', color=color_borde),
                right=Side(style='thin', color=color_borde),
                top=Side(style='thin', color=color_borde),
                bottom=Side(style='thin', color=color_borde)
            )
            
            # ============================================
            # SECCIÓN 1: AGREGAR ENCABEZADO CON LOGO (usando función auxiliar)
            # ============================================
            start_row = self._agregar_encabezado_excel(ws, sheet_name, incluir_logo=True)
            
            # Headers de la tabla
            headers = datos_tabla.get('headers', [])
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num)
                cell.value = str(header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Filas principales
            rows = datos_tabla.get('rows', [])
            rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
            
            current_row = start_row + 1
            for idx, row in enumerate(rows):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    if value is None:
                        cell.value = ""
                    elif isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, datetime):
                        cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        cell.value = str(value)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Agregar detalles si existen para esta fila Y se solicita mostrarlos
                if mostrar_detalles and idx < len(rows_con_detalles):
                    fila_detalles = rows_con_detalles[idx]
                    detalles = fila_detalles.get('detalles', [])
                    
                    for detalle in detalles:
                        detalle_datos = detalle.get('datos', [])
                        for datos_detalle in detalle_datos:
                            if isinstance(datos_detalle, dict):
                                titulo = datos_detalle.get('titulo', 'Detalle')
                                detail_rows = datos_detalle.get('rows', [])
                                detail_headers = datos_detalle.get('headers', [])
                                
                                # Fila de separación con título del detalle
                                current_row += 1
                                ws.cell(row=current_row, column=1, value=f"  {titulo}:")
                                ws.cell(row=current_row, column=1).font = subtitle_font
                                ws.cell(row=current_row, column=1).fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
                                
                                # Si hay headers de detalle, agregarlos
                                if detail_headers:
                                    current_row += 1
                                    for col_num, h in enumerate(detail_headers, 1):
                                        cell = ws.cell(row=current_row, column=col_num)
                                        cell.value = str(h)
                                        cell.font = detail_font
                                        cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
                                        cell.border = border
                                
                                # Filas de detalle
                                for detail_row in detail_rows:
                                    current_row += 1
                                    for col_num, value in enumerate(detail_row, 1):
                                        cell = ws.cell(row=current_row, column=col_num)
                                        if value is None:
                                            cell.value = ""
                                        else:
                                            cell.value = str(value)
                                        cell.font = detail_font
                                        cell.border = border
                
                current_row += 1
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                from openpyxl.utils import get_column_letter

                # Luego, en el método generar_excel_con_detalles(), reemplaza:
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar en buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise Exception(f"Error generando Excel con detalles: {str(e)}")

    def generar_excel_con_historial(self, tipo_reporte, datos_tabla, historial_membresias, filename="reporte.xlsx", sheet_name="Reporte"):
        """
        Genera un Excel profesional con historial de membresías incluido
        Args:
            tipo_reporte: Tipo de reporte (membresias, clientes, etc.)
            datos_tabla: Diccionario con headers, rows y rows_con_detalles
            historial_membresias: Diccionario con el historial de membresías por cliente_id
            filename: Nombre del archivo
            sheet_name: Nombre de la hoja
        """
        if not PANDAS_DISPONIBLE:
            raise Exception("Pandas/OpenPyXL no está disponible. Por favor instale: pip install pandas openpyxl")
        
        try:
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]
            
            # Convertir color hexadecimal a formato Excel (sin #)
            color_primario = self.config['color_primario'].replace('#', '')
            color_secundario = self.config['color_secundario'].replace('#', '')
            color_borde = self.COLORES['borde'].replace('#', '')
            color_header_bg = self.COLORES['header_bg'].replace('#', '')
            
            # Estilos
            header_fill = PatternFill(start_color=color_primario, 
                                       end_color=color_primario, 
                                       fill_type='solid')
            header_font = Font(color="FFFFFF", bold=True, size=11)
            historial_header_fill = PatternFill(start_color='6B7280', 
                                       end_color='6B7280', 
                                       fill_type='solid')
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(bold=True, size=11, color=color_primario)
            data_font = Font(size=10)
            detail_font = Font(size=9, italic=True)
            historial_font = Font(size=9)
            border = Border(
                left=Side(style='thin', color=color_borde),
                right=Side(style='thin', color=color_borde),
                top=Side(style='thin', color=color_borde),
                bottom=Side(style='thin', color=color_borde)
            )
            
            # ============================================
            # SECCIÓN 1: AGREGAR ENCABEZADO CON LOGO (usando función auxiliar)
            # ============================================
            start_row = self._agregar_encabezado_excel(ws, f"{sheet_name} con Historial", incluir_logo=True)
            
            # ============================================
            # SECCIÓN 2: ENCABEZADOS DE LA TABLA PRINCIPAL
            # ============================================
            headers = datos_tabla.get('headers', [])
            header_row = start_row
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = str(header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ============================================
            # SECCIÓN 3: DATOS PRINCIPALES CON HISTORIAL
            # ============================================
            rows = datos_tabla.get('rows', [])
            rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
            
            current_row = header_row + 1
            
            for idx, fila_con_detalles in enumerate(rows_con_detalles):
                row = fila_con_detalles.get('datos', [])
                cliente_id = fila_con_detalles.get('id')
                
                # Escribir datos principales
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    if value is None:
                        cell.value = ""
                    elif isinstance(value, (int, float)):
                        cell.value = value
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                    elif isinstance(value, datetime):
                        cell.value = value.strftime('%d/%m/%Y')
                    else:
                        cell.value = str(value)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Filas alternas con color
                if idx % 2 == 0:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.fill = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type='solid')
                
                current_row += 1
                
                # ============================================
                # SECCIÓN 4: HISTORIAL DE MEMBRESÍAS
                # ============================================
                if historial_membresias and str(cliente_id) in historial_membresias:
                    historial_data = historial_membresias[str(cliente_id)]
                    historial = historial_data.get('historial', [])
                    cliente_nombre = historial_data.get('cliente_nombre', '')
                    cliente_dni = historial_data.get('cliente_dni', '')
                    
                    if historial and len(historial) > 0:
                        # Fila de título del historial
                        ws.cell(row=current_row, column=1, value=f"  📋 HISTORIAL DE MEMBRESÍAS - {cliente_nombre} (DNI: {cliente_dni})")
                        ws.cell(row=current_row, column=1).font = subtitle_font
                        ws.cell(row=current_row, column=1).fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
                        ws.merge_cells(f'A{current_row}:I{current_row}')
                        current_row += 1
                        
                        # Headers del historial
                        historial_headers = ['Plan', 'Código', 'Fecha Inicio', 'Fecha Fin', 'Monto', 'Método Pago', 'Estado', 'Observaciones', 'Usuario']
                        for col_num, h in enumerate(historial_headers, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = h
                            cell.fill = historial_header_fill
                            cell.font = Font(color="FFFFFF", bold=True, size=9)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        current_row += 1
                        
                        # Filas del historial
                        for h_idx, h in enumerate(reversed(list(historial))):
                            ws.cell(row=current_row, column=1, value=h.get('plan_nombre', ''))
                            ws.cell(row=current_row, column=2, value=h.get('plan_codigo', ''))
                            ws.cell(row=current_row, column=3, value=h.get('fecha_inicio', ''))
                            ws.cell(row=current_row, column=4, value=h.get('fecha_fin', ''))
                            ws.cell(row=current_row, column=5, value=float(h.get('monto_pagado', 0)))
                            ws.cell(row=current_row, column=5).number_format = '#,##0.00'
                            ws.cell(row=current_row, column=6, value=h.get('metodo_pago', ''))
                            ws.cell(row=current_row, column=7, value=h.get('estado', ''))
                            ws.cell(row=current_row, column=8, value=h.get('observaciones', ''))
                            ws.cell(row=current_row, column=9, value=h.get('usuario_nombre', ''))
                            
                            for col_num in range(1, len(historial_headers) + 1):
                                cell = ws.cell(row=current_row, column=col_num)
                                cell.font = historial_font
                                cell.border = border
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                
                                # Color según estado - NUEVOS ESTADOS
                                estado = h.get('estado', '').lower()
                                if 'pagado' in estado:
                                    cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                                elif 'pendiente' in estado:
                                    cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                                elif 'activa' in estado:
                                    cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                                elif 'vencida' in estado or 'vencido' in estado:
                                    cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                                elif 'terminada' in estado or 'terminado' in estado:
                                    cell.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
                                elif h_idx % 2 == 0:
                                    cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
                            
                            current_row += 1
                        
                        # Fila separadora después del historial
                        current_row += 1
            
            # ============================================
            # SECCIÓN 5: PIE DE PÁGINA
            # ============================================
            footer_row = current_row + 1
            
            # Línea separadora
            ws.merge_cells(f'A{footer_row}:I{footer_row}')
            separator_cell = ws[f'A{footer_row}']
            separator_cell.border = Border(bottom=Side(style='medium', color=color_primario))
            
            # Información del pie
            footer_row += 1
            ws.merge_cells(f'A{footer_row}:I{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = f"Reporte generado - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            footer_cell.font = Font(size=9, italic=True, color=color_secundario)
            footer_cell.alignment = Alignment(horizontal='center')
            
            # Total de registros
            footer_row += 1
            ws.cell(row=footer_row, column=len(headers)-1, value="Total Registros:")
            ws.cell(row=footer_row, column=len(headers)-1).font = Font(bold=True)
            ws.cell(row=footer_row, column=len(headers)-1).alignment = Alignment(horizontal='right')
            
            ws.cell(row=footer_row, column=len(headers), value=len(rows))
            ws.cell(row=footer_row, column=len(headers)).font = Font(bold=True)
            
            # ============================================
            # SECCIÓN 6: AJUSTAR FORMATO
            # ============================================
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if isinstance(cell.value, datetime) or 'fecha' in str(cell.value).lower():
                                cell_length = 12
                            elif isinstance(cell.value, (int, float)):
                                cell_length = min(15, cell_length + 5)
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar paneles (encabezados fijos)
            ws.freeze_panes = ws[f'A{header_row + 1}']
            
            # Ajustar altura de filas
            for row in range(header_row, header_row + len(rows) + 1):
                ws.row_dimensions[row].height = 20
            
            # Guardar en buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise Exception(f"Error generando Excel con historial: {str(e)}")

    def generar_excel_con_entradas(self, tipo_reporte, datos_tabla, entradas_inventario, filename="reporte.xlsx", sheet_name="Reporte"):
        """
        Genera un Excel profesional con entradas de inventario incluido
        Args:
            tipo_reporte: Tipo de reporte (productos, etc.)
            datos_tabla: Diccionario con headers, rows y rows_con_detalles
            entradas_inventario: Diccionario con las entradas de inventario por producto_id
            filename: Nombre del archivo
            sheet_name: Nombre de la hoja
        """
        if not PANDAS_DISPONIBLE:
            raise Exception("Pandas/OpenPyXL no está disponible. Por favor instale: pip install pandas openpyxl")
        
        try:
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]
            
            # Convertir color hexadecimal a formato Excel (sin #)
            color_primario = self.config['color_primario'].replace('#', '')
            color_secundario = self.config['color_secundario'].replace('#', '')
            color_borde = self.COLORES['borde'].replace('#', '')
            color_header_bg = self.COLORES['header_bg'].replace('#', '')
            
            # Estilos
            header_fill = PatternFill(start_color=color_primario, 
                                       end_color=color_primario, 
                                       fill_type='solid')
            header_font = Font(color="FFFFFF", bold=True, size=11)
            entradas_header_fill = PatternFill(start_color='10B981', 
                                       end_color='10B981', 
                                       fill_type='solid')
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(bold=True, size=11, color=color_primario)
            data_font = Font(size=10)
            detail_font = Font(size=9, italic=True)
            entradas_font = Font(size=9)
            border = Border(
                left=Side(style='thin', color=color_borde),
                right=Side(style='thin', color=color_borde),
                top=Side(style='thin', color=color_borde),
                bottom=Side(style='thin', color=color_borde)
            )
            
            # ============================================
            # SECCIÓN 1: AGREGAR ENCABEZADO CON LOGO (usando función auxiliar)
            # ============================================
            start_row = self._agregar_encabezado_excel(ws, f"{sheet_name} con Entradas", incluir_logo=True)
            
            # ============================================
            # SECCIÓN 2: ENCABEZADOS DE LA TABLA PRINCIPAL
            # ============================================
            headers = datos_tabla.get('headers', [])
            header_row = start_row
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = str(header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ============================================
            # SECCIÓN 3: DATOS PRINCIPALES CON ENTRADAS
            # ============================================
            rows = datos_tabla.get('rows', [])
            rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
            
            current_row = header_row + 1
            
            for idx, fila_con_detalles in enumerate(rows_con_detalles):
                row = fila_con_detalles.get('datos', [])
                producto_id = fila_con_detalles.get('id')
                
                # Escribir datos principales
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    if value is None:
                        cell.value = ""
                    elif isinstance(value, (int, float)):
                        cell.value = value
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                    elif isinstance(value, datetime):
                        cell.value = value.strftime('%d/%m/%Y')
                    else:
                        cell.value = str(value)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Filas alternas con color
                if idx % 2 == 0:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.fill = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type='solid')
                
                current_row += 1
                
                # ============================================
                # SECCIÓN 4: ENTRADAS DE INVENTARIO
                # ============================================
                if entradas_inventario and producto_id and str(producto_id) in entradas_inventario:
                    entradas = entradas_inventario[str(producto_id)]
                    
                    if entradas and len(entradas) > 0:
                        # Fila de título de las entradas
                        # Buscar nombre del producto
                        nombre_producto = row[1] if len(row) > 1 else f"Producto #{producto_id}"
                        ws.cell(row=current_row, column=1, value=f"  📦 ENTRADAS DE INVENTARIO - {nombre_producto}")
                        ws.cell(row=current_row, column=1).font = subtitle_font
                        ws.cell(row=current_row, column=1).fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                        ws.merge_cells(f'A{current_row}:G{current_row}')
                        current_row += 1
                        
                        # Headers de las entradas
                        entradas_headers = ['Fecha', 'Cantidad', 'Costo Unit.', 'Subtotal', 'Proveedor', 'Usuario Registro', 'Observaciones']
                        for col_num, h in enumerate(entradas_headers, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = h
                            cell.fill = entradas_header_fill
                            cell.font = Font(color="FFFFFF", bold=True, size=9)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        current_row += 1
                        
                        # Filas de las entradas
                        total_unidades = 0
                        total_costo = 0
                        
                        for e_idx, entrada in enumerate(reversed(list(entradas))):
                            fecha = entrada.get('fecha_entrada', '')
                            if fecha and ' ' in fecha:
                                fecha = fecha.split(' ')[0]
                            
                            cantidad = int(entrada.get('cantidad', 0) or 0)
                            costo_unitario = float(entrada.get('costo_unitario', 0) or 0)
                            subtotal = cantidad * costo_unitario
                            
                            total_unidades += cantidad
                            total_costo += subtotal
                            
                            ws.cell(row=current_row, column=1, value=fecha)
                            ws.cell(row=current_row, column=2, value=cantidad)
                            ws.cell(row=current_row, column=2).number_format = '#,##0'
                            ws.cell(row=current_row, column=3, value=costo_unitario)
                            ws.cell(row=current_row, column=3).number_format = '#,##0.00'
                            ws.cell(row=current_row, column=4, value=subtotal)
                            ws.cell(row=current_row, column=4).number_format = '#,##0.00'
                            ws.cell(row=current_row, column=5, value=entrada.get('proveedor', ''))
                            ws.cell(row=current_row, column=6, value=entrada.get('usuario_registro', 'Sistema'))
                            ws.cell(row=current_row, column=7, value=entrada.get('observaciones', ''))
                            
                            for col_num in range(1, len(entradas_headers) + 1):
                                cell = ws.cell(row=current_row, column=col_num)
                                cell.font = entradas_font
                                cell.border = border
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                
                                # Color según cantidad
                                if e_idx % 2 == 0:
                                    cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
                            
                            current_row += 1
                        
                        # Fila de totales
                        ws.cell(row=current_row, column=1, value="TOTALES:")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
                        ws.cell(row=current_row, column=2, value=total_unidades)
                        ws.cell(row=current_row, column=2).font = Font(bold=True)
                        ws.cell(row=current_row, column=2).number_format = '#,##0'
                        ws.cell(row=current_row, column=4, value=total_costo)
                        ws.cell(row=current_row, column=4).font = Font(bold=True)
                        ws.cell(row=current_row, column=4).number_format = '#,##0.00'
                        
                        for col_num in range(1, len(entradas_headers) + 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
                            cell.border = border
                        
                        current_row += 1
                        
                        # Fila separadora después de las entradas
                        current_row += 1
            
            # ============================================
            # SECCIÓN 5: PIE DE PÁGINA
            # ============================================
            footer_row = current_row + 1
            
            # Línea separadora
            ws.merge_cells(f'A{footer_row}:G{footer_row}')
            separator_cell = ws[f'A{footer_row}']
            separator_cell.border = Border(bottom=Side(style='medium', color=color_primario))
            
            # Información del pie
            footer_row += 1
            ws.merge_cells(f'A{footer_row}:G{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = f"Reporte generado - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            footer_cell.font = Font(size=9, italic=True, color=color_secundario)
            footer_cell.alignment = Alignment(horizontal='center')
            
            # Total de registros
            footer_row += 1
            ws.cell(row=footer_row, column=len(headers)-1, value="Total Registros:")
            ws.cell(row=footer_row, column=len(headers)-1).font = Font(bold=True)
            ws.cell(row=footer_row, column=len(headers)-1).alignment = Alignment(horizontal='right')
            
            ws.cell(row=footer_row, column=len(headers), value=len(rows))
            ws.cell(row=footer_row, column=len(headers)).font = Font(bold=True)
            
            # ============================================
            # SECCIÓN 6: AJUSTAR FORMATO
            # ============================================
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if isinstance(cell.value, datetime) or 'fecha' in str(cell.value).lower():
                                cell_length = 12
                            elif isinstance(cell.value, (int, float)):
                                cell_length = min(15, cell_length + 5)
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar paneles (encabezados fijos)
            ws.freeze_panes = ws[f'A{header_row + 1}']
            
            # Ajustar altura de filas
            for row in range(header_row, header_row + len(rows) + 1):
                ws.row_dimensions[row].height = 20
            
            # Guardar en buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise Exception(f"Error generando Excel con entradas: {str(e)}")

    def generar_excel_con_detalles_ventas(self, tipo_reporte, datos_tabla, detalles_ventas, filename="reporte.xlsx", sheet_name="Reporte"):
        """
        Genera un Excel profesional con detalles de ventas incluido
        Args:
            tipo_reporte: Tipo de reporte (ventas, etc.)
            datos_tabla: Diccionario con headers, rows y rows_con_detalles
            detalles_ventas: Diccionario con los detalles de ventas por venta_id
            filename: Nombre del archivo
            sheet_name: Nombre de la hoja
        """
        if not PANDAS_DISPONIBLE:
            raise Exception("Pandas/OpenPyXL no está disponible. Por favor instale: pip install pandas openpyxl")
        
        try:
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]
            
            # Convertir color hexadecimal a formato Excel (sin #)
            color_primario = self.config['color_primario'].replace('#', '')
            color_secundario = self.config['color_secundario'].replace('#', '')
            color_borde = self.COLORES['borde'].replace('#', '')
            color_header_bg = self.COLORES['header_bg'].replace('#', '')
            
            # Estilos
            header_fill = PatternFill(start_color=color_primario, 
                                       end_color=color_primario, 
                                       fill_type='solid')
            header_font = Font(color="FFFFFF", bold=True, size=11)
            detalles_header_fill = PatternFill(start_color='3B82F6', 
                                       end_color='3B82F6', 
                                       fill_type='solid')
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(bold=True, size=11, color=color_primario)
            data_font = Font(size=10)
            detail_font = Font(size=9, italic=True)
            detalles_font = Font(size=9)
            border = Border(
                left=Side(style='thin', color=color_borde),
                right=Side(style='thin', color=color_borde),
                top=Side(style='thin', color=color_borde),
                bottom=Side(style='thin', color=color_borde)
            )
            
            # ============================================
            # SECCIÓN 1: AGREGAR ENCABEZADO CON LOGO (usando función auxiliar)
            # ============================================
            start_row = self._agregar_encabezado_excel(ws, f"{sheet_name} con Detalles", incluir_logo=True)
            
            # ============================================
            # SECCIÓN 2: ENCABEZADOS DE LA TABLA PRINCIPAL
            # ============================================
            headers = datos_tabla.get('headers', [])
            header_row = start_row
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = str(header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ============================================
            # SECCIÓN 3: DATOS PRINCIPALES CON DETALLES
            # ============================================
            rows = datos_tabla.get('rows', [])
            rows_con_detalles = datos_tabla.get('rows_con_detalles', [])
            
            current_row = header_row + 1
            total_general = 0
            
            for idx, fila_con_detalles in enumerate(rows_con_detalles):
                row = fila_con_detalles.get('datos', [])
                venta_id = fila_con_detalles.get('id')
                
                # Escribir datos principales
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    if value is None:
                        cell.value = ""
                    elif isinstance(value, (int, float)):
                        cell.value = value
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                    elif isinstance(value, datetime):
                        cell.value = value.strftime('%d/%m/%Y')
                    else:
                        cell.value = str(value)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Calcular total para esta venta
                if venta_id and str(venta_id) in detalles_ventas:
                    detalles = detalles_ventas[str(venta_id)]
                    total_venta = sum(d.get('subtotal', 0) for d in detalles)
                    total_general += total_venta
                
                # Filas alternas con color
                if idx % 2 == 0:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.fill = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type='solid')
                
                current_row += 1
                
                # ============================================
                # SECCIÓN 4: DETALLES DE LA VENTA
                # ============================================
                if detalles_ventas and venta_id and str(venta_id) in detalles_ventas:
                    detalles = detalles_ventas[str(venta_id)]
                    
                    if detalles and len(detalles) > 0:
                        # Fila de título de los detalles
                        ws.cell(row=current_row, column=1, value=f"  🛒 DETALLE DE VENTA #{venta_id}")
                        ws.cell(row=current_row, column=1).font = subtitle_font
                        ws.cell(row=current_row, column=1).fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
                        ws.merge_cells(f'A{current_row}:G{current_row}')
                        current_row += 1
                        
                        # Headers de los detalles
                        detalles_headers = ['Producto', 'Categoría', 'Cantidad', 'Precio Unit.', 'Subtotal']
                        for col_num, h in enumerate(detalles_headers, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = h
                            cell.fill = detalles_header_fill
                            cell.font = Font(color="FFFFFF", bold=True, size=9)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Agregar celdas vacías para completar la fila
                        for col_num in range(len(detalles_headers) + 1, len(headers) + 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.fill = detalles_header_fill
                            cell.border = border
                        
                        current_row += 1
                        
                        # Filas de los detalles
                        total_detalle = 0
                        for d_idx, detalle in enumerate(detalles):
                            cantidad = detalle.get('cantidad', 0)
                            precio_unitario = detalle.get('precio_unitario', 0)
                            subtotal = detalle.get('subtotal', 0)
                            total_detalle += subtotal
                            
                            ws.cell(row=current_row, column=1, value=detalle.get('producto_nombre', detalle.get('nombre', '')))
                            ws.cell(row=current_row, column=2, value=detalle.get('categoria', ''))
                            ws.cell(row=current_row, column=3, value=cantidad)
                            ws.cell(row=current_row, column=3).number_format = '#,##0'
                            ws.cell(row=current_row, column=4, value=precio_unitario)
                            ws.cell(row=current_row, column=4).number_format = '#,##0.00'
                            ws.cell(row=current_row, column=5, value=subtotal)
                            ws.cell(row=current_row, column=5).number_format = '#,##0.00'
                            
                            for col_num in range(1, len(headers) + 1):
                                cell = ws.cell(row=current_row, column=col_num)
                                cell.font = detalles_font
                                cell.border = border
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                
                                # Color según índice
                                if d_idx % 2 == 0:
                                    cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
                            
                            current_row += 1
                        
                        # Fila de total del detalle
                        ws.cell(row=current_row, column=1, value="TOTAL VENTA:")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
                        
                        # Celdas vacías
                        for col_num in range(2, 5):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.border = border
                        
                        ws.cell(row=current_row, column=5, value=total_detalle)
                        ws.cell(row=current_row, column=5).font = Font(bold=True)
                        ws.cell(row=current_row, column=5).number_format = '#,##0.00'
                        
                        for col_num in range(1, len(headers) + 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
                            cell.border = border
                        
                        current_row += 1
                        
                        # Fila separadora después de los detalles
                        current_row += 1
            
            # ============================================
            # SECCIÓN 5: PIE DE PÁGINA
            # ============================================
            footer_row = current_row + 1
            
            # Línea separadora
            ws.merge_cells(f'A{footer_row}:G{footer_row}')
            separator_cell = ws[f'A{footer_row}']
            separator_cell.border = Border(bottom=Side(style='medium', color=color_primario))
            
            # Información del pie
            footer_row += 1
            ws.merge_cells(f'A{footer_row}:G{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = f"Reporte generado - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            footer_cell.font = Font(size=9, italic=True, color=color_secundario)
            footer_cell.alignment = Alignment(horizontal='center')
            
            # Total de registros y monto
            footer_row += 1
            ws.cell(row=footer_row, column=len(headers)-2, value="Total Registros:")
            ws.cell(row=footer_row, column=len(headers)-2).font = Font(bold=True)
            ws.cell(row=footer_row, column=len(headers)-2).alignment = Alignment(horizontal='right')
            
            ws.cell(row=footer_row, column=len(headers)-1, value=len(rows))
            ws.cell(row=footer_row, column=len(headers)-1).font = Font(bold=True)
            
            footer_row += 1
            ws.cell(row=footer_row, column=len(headers)-2, value="Total General:")
            ws.cell(row=footer_row, column=len(headers)-2).font = Font(bold=True)
            ws.cell(row=footer_row, column=len(headers)-2).alignment = Alignment(horizontal='right')
            
            ws.cell(row=footer_row, column=len(headers)-1, value=total_general)
            ws.cell(row=footer_row, column=len(headers)-1).font = Font(bold=True)
            ws.cell(row=footer_row, column=len(headers)-1).number_format = '#,##0.00'
            
            # ============================================
            # SECCIÓN 6: AJUSTAR FORMATO
            # ============================================
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if isinstance(cell.value, datetime) or 'fecha' in str(cell.value).lower():
                                cell_length = 12
                            elif isinstance(cell.value, (int, float)):
                                cell_length = min(15, cell_length + 5)
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar paneles (encabezados fijos)
            ws.freeze_panes = ws[f'A{header_row + 1}']
            
            # Ajustar altura de filas
            for row in range(header_row, header_row + len(rows) + 1):
                ws.row_dimensions[row].height = 20
            
            # Guardar en buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise Exception(f"Error generando Excel con detalles de ventas: {str(e)}")


# ============================================
# FUNCIÓN AUXILIAR PARA OBTENER DATOS
# ============================================

def obtener_datos_reporte(tipo_reporte, db_path=None, fecha_inicio=None, fecha_fin=None):
    """Función auxiliar para obtener datos de reportes"""
    if db_path is None:
        db_path = 'sistema.db'
    
    resultado = {'tipo': tipo_reporte, 'reporte': [], 'stats': {}}
    
    try:
        if tipo_reporte == 'clientes':
            query = '''
                SELECT c.*, p.nombre as plan_nombre, p.precio as plan_precio
                FROM clientes c
                LEFT JOIN planes_membresia p ON c.plan_id = p.id
                WHERE c.activo = 1
                ORDER BY c.fecha_inicio DESC
            '''
            resultado['reporte'] = execute_query(query, fetch=True)
        
        elif tipo_reporte == 'ventas':
            fecha_query = ""
            params = []
            if fecha_inicio and fecha_fin:
                if is_mysql():
                    fecha_query = "AND DATE(v.fecha_venta) BETWEEN %s AND %s"
                else:
                    fecha_query = "AND DATE(v.fecha_venta) BETWEEN ? AND ?"
                params = [fecha_inicio, fecha_fin]
            
            query = f'''
                SELECT v.*, COUNT(dv.id) as productos_count
                FROM ventas v
                LEFT JOIN detalle_ventas dv ON v.id = dv.venta_id
                WHERE (v.estado != 'eliminado' OR v.estado IS NULL) {fecha_query}
                GROUP BY v.id
                ORDER BY v.fecha_venta DESC
            '''
            resultado['reporte'] = execute_query(query, params, fetch=True)
        
        elif tipo_reporte == 'productos':
            query = '''
                SELECT * FROM productos 
                WHERE activo = 1
                ORDER BY nombre
            '''
            resultado['reporte'] = execute_query(query, fetch=True)

        elif tipo_reporte == 'promociones':
            query = '''
                SELECT p.*, pm.nombre as plan_nombre
                FROM promociones p
                LEFT JOIN planes_membresia pm ON p.plan_id = pm.id
                ORDER BY p.fecha_inicio DESC
            '''
            resultado['reporte'] = execute_query(query, fetch=True)
    
    except Exception as e:
        print(f"Error obteniendo datos de reporte: {e}")
    
    return resultado